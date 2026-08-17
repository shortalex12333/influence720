"""
Influence 720 — Part I: export the 2026-08-16 coverage-catalog research (Excel-first: the master
workbook was enriched FIRST via coverage_research_v2.json -> foundation_v2.COVERAGE_CATALOG ->
build_master_excel.py's CovCatalog_<Vertical> sheets; this script reads THAT workbook back out,
never the original Python objects or the research JSON directly — same discipline
regenerate_from_master.py already holds itself to for every other dataset in this project).

Two outputs, both byte-traceable to Influence720_MasterOntology.xlsx:

  coverage-corpus.json  — the TEMPLATE-level catalog: coverage types, baseline limits, statutory
                           triggers, sub-limit exclusions, cascade rules, per vertical. Sourced
                           from the 6 CovCatalog_<Vertical> sheets.

  epic-staged-ledger.json — the REAL per-account coverage/bound-status ledger for all 150+
                           accounts, sourced from Coverages + Overlay_BoundPolicy + Policies +
                           Assets + UBO_Principals. Framed as an Epic policy-management-system
                           export — STAGED/mock, matching this project's existing "Salesforce/Epic
                           mock overlays" framing already used in 05-build/data/overlays.js
                           ("// Influence 720 — Salesforce/Epic mock overlays. Regenerated FROM
                           Influence720_MasterOntology.xlsx.").

Run AFTER the full 9-step pipeline (flagship_and_allocate_v2.py through regenerate_from_master.py)
so the master workbook actually reflects the current catalog + current per-account dataset.
"""
import json
import openpyxl

from regenerate_from_master import MASTER_XLSX, read_sheet, SHEET_SUFFIX_TO_VERTICAL

DATA_OUT_DIR = "/Users/celeste7/Documents/INFLUENCE/05-build/data"

COV_CATALOG_SHEET_PREFIX = "CovCatalog_"


def _parse_json_col(val):
    if not val:
        return []
    if isinstance(val, (list, dict)):
        return val
    return json.loads(val)


def build_coverage_corpus(wb):
    """Template-level catalog, one row per coverage_type_id per vertical — straight read-back
    of the CovCatalog_<Vertical> sheets build_master_excel.py wrote from foundation_v2.COVERAGE_CATALOG."""
    verticals = {}
    total_rows = 0
    total_new = 0
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(COV_CATALOG_SHEET_PREFIX):
            continue
        suffix = sheet_name[len(COV_CATALOG_SHEET_PREFIX):]
        vertical_key = SHEET_SUFFIX_TO_VERTICAL.get(suffix, suffix.lower())
        rows = []
        for r in read_sheet(wb, sheet_name):
            entry = {
                "coverage_type_id": r["coverage_type_id"],
                "label": r["label"],
                "trigger_type": r["trigger_type"],
                "line_of_business": r["line_of_business"],
                "is_new": bool(r.get("is_new")),
                "required_baseline_limit": r.get("required_baseline_limit") or "",
                "baseline_limit_usd": r.get("baseline_limit_usd"),
                "baseline_limit_by_tier": _parse_json_col(r.get("baseline_limit_by_tier_json")) or None,
                "statutory_triggers": _parse_json_col(r.get("statutory_triggers_json")),
                "sub_limit_exclusions": _parse_json_col(r.get("sub_limit_exclusions_json")),
                "cascade_rules": _parse_json_col(r.get("cascade_rules_json")),
                "note": r.get("note") or "",
            }
            rows.append(entry)
            total_rows += 1
            if entry["is_new"]:
                total_new += 1
        verticals[vertical_key] = rows

    return {
        "_source": "Influence720_MasterOntology.xlsx — CovCatalog_<Vertical> sheets (regenerated FROM the "
                    "master workbook by gen_coverage_corpus.py; never hand-authored separately from it).",
        "_generated_by": "05-build/generator/build_v2/gen_coverage_corpus.py",
        "verticals": verticals,
        "total_coverage_types": total_rows,
        "new_coverage_types_2026_08_16": total_new,
    }


def build_epic_staged_ledger(wb):
    """Real per-account coverage/bound-status ledger, joined from Coverages + Overlay_BoundPolicy
    + Policies + Assets + UBO_Principals — every row here traces to a real per-account Coverage
    node in the master workbook, not a template row (see coverage-corpus.json for the template)."""
    principals_by_id = {r["id"]: r for r in read_sheet(wb, "UBO_Principals")}
    assets_by_id = {r["id"]: r for r in read_sheet(wb, "Assets")}
    policies_by_id = {r["id"]: r for r in read_sheet(wb, "Policies")}
    overlay_by_cov_id = {r["coverage_id"]: r for r in read_sheet(wb, "Overlay_BoundPolicy")}

    accounts = {}
    unmatched_coverages = 0
    for cov in read_sheet(wb, "Coverages"):
        asset = assets_by_id.get(cov["parent_asset_ref"])
        account_id = asset.get("owner_ref") if asset else None
        if account_id is None or account_id not in principals_by_id:
            unmatched_coverages += 1
            continue
        overlay = overlay_by_cov_id.get(cov["id"], {})
        policy_id = overlay.get("policy_id") or cov.get("policy_id") or None
        policy = policies_by_id.get(policy_id) if policy_id else None

        acc = accounts.setdefault(account_id, {
            "account_id": account_id,
            "principal_label": principals_by_id[account_id]["label"],
            "vertical": principals_by_id[account_id]["vertical"],
            "producer_id": principals_by_id[account_id].get("producer_id") or "",
            "producer_name": principals_by_id[account_id].get("producer_name") or "",
            "coverages": [],
        })
        acc["coverages"].append({
            "coverage_id": cov["id"],
            "label": cov["label"],
            "vertical": cov["vertical"],
            "coverage_type_id": cov["coverage_type_id"],
            "trigger_type": cov.get("trigger_type"),
            "line_of_business": cov.get("line_of_business"),
            "parent_asset_ref": cov["parent_asset_ref"],
            "parent_asset_label": asset.get("label") if asset else None,
            "status": overlay.get("status"),
            "bound_premium": overlay.get("bound_premium"),
            "bound_limit_usd": overlay.get("bound_limit_usd"),
            "carrier": overlay.get("carrier"),
            "gwp_estimate": overlay.get("gwp_estimate"),
            "gwp_method": overlay.get("gwp_method"),
            "gwp_confidence": overlay.get("gwp_confidence"),
            "gwp_peer_n": overlay.get("gwp_peer_n"),
            "rate_trend_source": overlay.get("rate_trend_source"),
            "unbound_reason": overlay.get("unbound_reason"),
            "evidence_note": overlay.get("evidence_note"),
            "expiry_date": overlay.get("expiry_date"),
            "producer_of_record": overlay.get("producer_of_record"),
            "policy_id": policy_id,
            "policy_number": policy.get("policy_number") if policy else None,
        })

    status_counts = {}
    for acc in accounts.values():
        for c in acc["coverages"]:
            status_counts[c["status"]] = status_counts.get(c["status"], 0) + 1

    return {
        "_epic_export_note": (
            "STAGED — presented as if sourced from an Epic policy management system export. Entirely "
            "generated from Influence720_MasterOntology.xlsx (Coverages + Overlay_BoundPolicy + Policies + "
            "Assets + UBO_Principals sheets), consistent with this project's existing 'Salesforce/Epic mock "
            "overlays' framing already used in 05-build/data/overlays.js. No real Epic system was accessed; "
            "every row here traces back to the master workbook, never independently authored."
        ),
        "_source": "Influence720_MasterOntology.xlsx — Coverages, Overlay_BoundPolicy, Policies, Assets, UBO_Principals sheets.",
        "_generated_by": "05-build/generator/build_v2/gen_coverage_corpus.py",
        "accounts": list(accounts.values()),
        "total_accounts": len(accounts),
        "total_coverage_rows": sum(len(a["coverages"]) for a in accounts.values()),
        "status_breakdown": status_counts,
        "unmatched_coverage_rows": unmatched_coverages,
    }


def build_gwp_rate_trend(wb):
    """GWP_RateTrend_Table sheet, keyed by coverage_type_id — the same real-sourced (IUMI Stats
    Report 2025 / Marsh GIMI) annual rate-trend figures GWP_Pricing_Methodology already documents
    for the pricing engine, exported here so a dashboard can cite them directly instead of
    re-deriving them. [NO SOURCE FOUND — ESTIMATE] entries are carried through verbatim, never
    silently dropped or relabeled as sourced."""
    rows = list(wb["GWP_RateTrend_Table"].iter_rows(values_only=True))
    # row 0 is a note row (col1 only), row 1 is the header
    header_idx = 1 if len(rows) > 1 and all(v is None for v in rows[0][1:]) else 0
    out = {}
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        out[r[0]] = {"annual_trend_pct": r[1], "source": r[2]}
    return out


def build_gwp_pricing_methodology(wb):
    rows = list(wb["GWP_Pricing_Methodology"].iter_rows(values_only=True))
    header = rows[0]
    return [{header[i]: r[i] for i in range(len(header))} for r in rows[1:] if r[0] is not None]


def emit_js(corpus, rate_trend, pricing_methodology):
    """A script-includable sibling of coverage-corpus.json — the JSON file has no `const X =`
    wrapper, so a dashboard cannot <script src> it directly (and this project's dashboards are
    opened as much via file:// as via a server, where fetch() of a local JSON file is blocked by
    CORS). Same read-back-from-the-workbook discipline as regenerate_from_master.py's own js()
    helper; not a duplicate data source, just a loadable transport for the identical content."""
    with open(f"{DATA_OUT_DIR}/coverage-corpus.js", "w") as f:
        f.write("// Influence 720 — coverage catalog + GWP rate-trend table. Script-loadable twin of "
                "coverage-corpus.json, regenerated FROM Influence720_MasterOntology.xlsx by gen_coverage_corpus.py.\n")
        f.write(f"const COVERAGE_CORPUS = {json.dumps(corpus, indent=2, default=str)};\n")
        f.write(f"const GWP_RATE_TREND = {json.dumps(rate_trend, indent=2, default=str)};\n")
        f.write(f"const GWP_PRICING_METHODOLOGY = {json.dumps(pricing_methodology, indent=2, default=str)};\n")
    print(f"coverage-corpus.js: {len(rate_trend)} rate-trend rows, {len(pricing_methodology)} pricing-methodology rows")


def main():
    import os
    os.makedirs(DATA_OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)

    corpus = build_coverage_corpus(wb)
    with open(f"{DATA_OUT_DIR}/coverage-corpus.json", "w") as f:
        json.dump(corpus, f, indent=2, default=str)
    print(f"coverage-corpus.json: {corpus['total_coverage_types']} coverage types across "
          f"{len(corpus['verticals'])} verticals ({corpus['new_coverage_types_2026_08_16']} new)")

    ledger = build_epic_staged_ledger(wb)
    with open(f"{DATA_OUT_DIR}/epic-staged-ledger.json", "w") as f:
        json.dump(ledger, f, indent=2, default=str)
    print(f"epic-staged-ledger.json: {ledger['total_accounts']} accounts, "
          f"{ledger['total_coverage_rows']} coverage rows, status breakdown: {ledger['status_breakdown']}")
    if ledger["unmatched_coverage_rows"]:
        print(f"  WARNING: {ledger['unmatched_coverage_rows']} coverage rows could not be matched to an account")

    rate_trend = build_gwp_rate_trend(wb)
    pricing_methodology = build_gwp_pricing_methodology(wb)
    emit_js(corpus, rate_trend, pricing_methodology)


if __name__ == "__main__":
    main()
