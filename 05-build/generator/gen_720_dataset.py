import json

OUT_DIR = "/Users/celeste7/Documents/INFLUENCE/05-build/data"

# ============================================================
# ROLE / PRINCIPAL NODES (representative subset, not the full 187 —
# reuses real 360 ids for yacht overlap, new ids for other verticals)
# ============================================================
nodes = []

def role(id_, label, node_type, vertical, ring, domain, coi_type, desc, coi_weight=7, coi_freq=6):
    nodes.append({
        "id": id_, "label": label, "node_type": node_type, "vertical": vertical,
        "ring": ring, "domain": domain, "coi_type": coi_type, "coi_weight": coi_weight,
        "coi_freq": coi_freq, "desc": desc,
    })

def asset(id_, label, vertical, owner_ref, asset_class, size_class, required_coverage_ids):
    nodes.append({
        "id": id_, "label": label, "node_type": "asset", "vertical": vertical,
        "owner_ref": owner_ref, "asset_class": asset_class, "size_class": size_class,
        "required_coverage_ids": required_coverage_ids,
    })

def coverage(id_, label, vertical, coverage_type_id, parent_asset_ref):
    nodes.append({
        "id": id_, "label": label, "node_type": "coverage", "vertical": vertical,
        "coverage_type_id": coverage_type_id, "parent_asset_ref": parent_asset_ref,
    })

# ---- Principals (UBOs) — random staged names ----
role("PRIN.MARCHETTI", "Marchetti Family Trust", "principal", "yacht", 0, 1, "signatory",
     "Apex — UBO, Whisper/Gulfstream ownership structure.", 10, 3)
role("PRIN.MERIDIAN", "Meridian Family Office", "principal", "yacht", 0, 1, "signatory",
     "Apex — UBO, M/Y Solstice.", 10, 3)
role("PRIN.HARTWELL", "Hartwell Holdings", "principal", "yacht", 0, 1, "signatory",
     "Apex — UBO, M/Y Aurelia.", 10, 3)
role("PRIN.VOSS", "Voss Family Trust", "principal", "yacht", 0, 1, "signatory",
     "Apex — UBO, M/Y Halcyon + Aspen chalet.", 10, 3)
role("PRIN.CASTELLANE", "Castellane Trust", "principal", "aviation", 0, 1, "signatory",
     "Apex — UBO, Gulfstream G550 + hypercar collection + personal risk program.", 10, 3)
role("PRIN.BELLWEATHER", "Bellweather Family Office", "principal", "aviation", 0, 1, "signatory",
     "Apex — UBO, Sikorsky S-76.", 10, 3)
role("PRIN.OKONKWO", "Okonkwo Family Trust", "principal", "yacht", 0, 1, "signatory",
     "Apex — UBO, M/Y Meridian II (lapsed/lost account).", 10, 3)

# ---- Yacht-vertical roles (reusing real 360 ids where they exist) ----
role("OWN.7", "Sarah Kellerman — General Counsel", "role", "yacht", 1, 1, "signatory",
     "GC for the Marchetti Family Office. Dual-hat: also the Aviation program's risk buyer of record (see cross-pollination edge).", 9, 6)
role("OWN.9", "Elena Marchetti — SPV Director", "role", "yacht", 1, 1, "signatory",
     "Yacht-owning SPV Director, Marchetti structure.", 8, 7)
role("OWN.12", "Richard Voss — Trustee", "role", "yacht", 1, 1, "signatory",
     "Trustee, Voss Family Trust — controls both the yacht SPV and the Aspen real-estate holding.", 9, 6)
role("MCR.2.WHISPER", "Capt. James Doyle", "role", "yacht", 1, 7, "signatory",
     "Captain, M/Y Whisper.", 10, 10)
role("OWN.7.MERIDIAN", "Diane Foster — Trustee", "role", "yacht", 1, 1, "signatory",
     "Trustee, Meridian Family Office. Ring-1 signatory on M/Y Solstice — comm frequency has decayed to zero over the last 8 months.", 9, 6)
role("MCR.2.SOLSTICE", "Capt. Owen Rees", "role", "yacht", 1, 7, "signatory",
     "Captain, M/Y Solstice.", 9, 10)
role("MCR.3.SOLSTICE", "Chief Eng. Priya Deshmukh", "role", "yacht", 1, 7, "signatory",
     "Chief Engineer, M/Y Solstice.", 8, 9)
role("OWN.9.HARTWELL", "Nigel Ashworth — SPV Director", "role", "yacht", 1, 1, "signatory",
     "SPV Director, Hartwell Holdings.", 8, 7)
role("MCR.2.AURELIA", "Capt. Lena Brandt", "role", "yacht", 1, 7, "signatory",
     "Captain, M/Y Aurelia — high-frequency contact, no purchasing authority.", 9, 10)
role("MCR.4.AURELIA", "ETO Sam Whitfield", "role", "yacht", 2, 7, "conditional",
     "ETO, M/Y Aurelia — high-frequency contact, no purchasing authority.", 5, 8)
role("MCR.6.AURELIA", "Bosun Theo Marsh", "role", "yacht", 3, 7, "operational",
     "Bosun, M/Y Aurelia — high-frequency contact, no purchasing authority.", 3, 7)
role("OWN.9.VOSS", "Isabelle Voss — SPV Director", "role", "yacht", 1, 1, "signatory",
     "Yacht SPV Director, also sits on the family trust's real-estate holding entity.", 8, 7)
role("OWN.9.OKONKWO", "Chidi Okonkwo — SPV Director", "role", "yacht", 1, 1, "signatory",
     "SPV Director, Okonkwo Family Trust — account fully lapsed 14 months ago.", 7, 2)

# ---- Aviation-vertical roles ----
role("AVI.RISK.CASTELLANE", "Marcus Delacroix — Family Office Risk Manager", "role", "aviation", 1, 4, "signatory",
     "Aviation program risk buyer, Castellane Trust.", 9, 6)
role("AVI.PILOT.CASTELLANE", "Chief Pilot Renata Solis", "role", "aviation", 1, 2, "conditional",
     "Chief Pilot, Castellane G550.", 6, 8)
role("AVI.RISK.BELLWEATHER", "Howard Ngata — Family Office Risk Manager", "role", "aviation", 1, 4, "signatory",
     "Aviation program risk buyer, Bellweather Family Office.", 9, 6)
role("AVI.PILOT.BELLWEATHER", "Chief Pilot Dana Okafor", "role", "aviation", 1, 2, "conditional",
     "Chief Pilot, Bellweather S-76.", 6, 8)

# ---- Real Estate-vertical roles ----
role("RE.PM.VOSS", "Fiona Aldridge — Property Manager", "role", "real_estate", 1, 3, "conditional",
     "Property Manager, Aspen chalet, Voss Family Trust.", 6, 6)

# ---- Automotive / Personal-vertical roles ----
role("AUTO.MGR.CASTELLANE", "Devon Pratt — Collection Manager", "role", "auto", 1, 1, "conditional",
     "Manages the 12-car hypercar collection, Castellane Trust.", 5, 5)
role("PER.EA.CASTELLANE", "Odette Vance — Executive Assistant", "role", "personal", 2, 1, "conditional",
     "EA to the Castellane principal — gatekeeper for personal-risk conversations.", 6, 7)

# ============================================================
# ASSETS + COVERAGES
# ============================================================
asset("AST.WHISPER.YACHT", "M/Y Whisper, 70m", "yacht", "PRIN.MARCHETTI", "70m Superyacht", "S6",
      ["COV.WHISPER.HULL", "COV.WHISPER.PI", "COV.WHISPER.CREWMED", "COV.WHISPER.FINEART"])
coverage("COV.WHISPER.HULL", "Hull & Machinery", "yacht", "hull_machinery", "AST.WHISPER.YACHT")
coverage("COV.WHISPER.PI", "P&I", "yacht", "pi", "AST.WHISPER.YACHT")
coverage("COV.WHISPER.CREWMED", "Crew Medical (MLC)", "yacht", "crew_medical_mlc", "AST.WHISPER.YACHT")
coverage("COV.WHISPER.FINEART", "Fine Art", "yacht", "fine_art", "AST.WHISPER.YACHT")

asset("AST.WHISPER.JET", "Gulfstream G650", "aviation", "PRIN.MARCHETTI", "Long-range business jet", "N/A",
      ["COV.WHISPERJET.HULL", "COV.WHISPERJET.LIAB", "COV.WHISPERJET.CREWACC", "COV.WHISPERJET.WAR"])
coverage("COV.WHISPERJET.HULL", "Aviation Hull", "aviation", "aviation_hull", "AST.WHISPER.JET")
coverage("COV.WHISPERJET.LIAB", "Aviation Liability", "aviation", "aviation_liability", "AST.WHISPER.JET")
coverage("COV.WHISPERJET.CREWACC", "Crew Accident", "aviation", "crew_accident", "AST.WHISPER.JET")
coverage("COV.WHISPERJET.WAR", "War Risk", "aviation", "war_risk", "AST.WHISPER.JET")

asset("AST.SOLSTICE.YACHT", "M/Y Solstice, 55m", "yacht", "PRIN.MERIDIAN", "55m Superyacht", "S5",
      ["COV.SOLSTICE.HULL", "COV.SOLSTICE.PI", "COV.SOLSTICE.CREWMED"])
coverage("COV.SOLSTICE.HULL", "Hull & Machinery", "yacht", "hull_machinery", "AST.SOLSTICE.YACHT")
coverage("COV.SOLSTICE.PI", "P&I", "yacht", "pi", "AST.SOLSTICE.YACHT")
coverage("COV.SOLSTICE.CREWMED", "Crew Medical (MLC)", "yacht", "crew_medical_mlc", "AST.SOLSTICE.YACHT")

asset("AST.AURELIA.YACHT", "M/Y Aurelia, 62m", "yacht", "PRIN.HARTWELL", "62m Superyacht", "S5",
      ["COV.AURELIA.HULL", "COV.AURELIA.PI", "COV.AURELIA.CREWMED", "COV.AURELIA.FINEART"])
coverage("COV.AURELIA.HULL", "Hull & Machinery", "yacht", "hull_machinery", "AST.AURELIA.YACHT")
coverage("COV.AURELIA.PI", "P&I", "yacht", "pi", "AST.AURELIA.YACHT")
coverage("COV.AURELIA.CREWMED", "Crew Medical (MLC)", "yacht", "crew_medical_mlc", "AST.AURELIA.YACHT")
coverage("COV.AURELIA.FINEART", "Fine Art", "yacht", "fine_art", "AST.AURELIA.YACHT")

asset("AST.HALCYON.YACHT", "M/Y Halcyon, 45m", "yacht", "PRIN.VOSS", "45m Superyacht", "S4",
      ["COV.HALCYON.HULL", "COV.HALCYON.PI"])
coverage("COV.HALCYON.HULL", "Hull & Machinery", "yacht", "hull_machinery", "AST.HALCYON.YACHT")
coverage("COV.HALCYON.PI", "P&I", "yacht", "pi", "AST.HALCYON.YACHT")

asset("AST.VOSS.CHALET", "Aspen Ski Chalet", "real_estate", "PRIN.VOSS", "High-Value Residential", "N/A",
      ["COV.CHALET.WINDFLOOD", "COV.CHALET.PROPERTY", "COV.CHALET.LIAB"])
coverage("COV.CHALET.WINDFLOOD", "Coastal Wind/Flood", "real_estate", "coastal_wind_flood", "AST.VOSS.CHALET")
coverage("COV.CHALET.PROPERTY", "High-Value Property", "real_estate", "high_value_property", "AST.VOSS.CHALET")
coverage("COV.CHALET.LIAB", "Premises Liability", "real_estate", "premises_liability", "AST.VOSS.CHALET")

asset("AST.CASTELLANE.JET", "Gulfstream G550", "aviation", "PRIN.CASTELLANE", "Long-range business jet", "N/A",
      ["COV.CASTJET.HULL", "COV.CASTJET.LIAB", "COV.CASTJET.CREWACC"])
coverage("COV.CASTJET.HULL", "Aviation Hull", "aviation", "aviation_hull", "AST.CASTELLANE.JET")
coverage("COV.CASTJET.LIAB", "Aviation Liability", "aviation", "aviation_liability", "AST.CASTELLANE.JET")
coverage("COV.CASTJET.CREWACC", "Crew Accident", "aviation", "crew_accident", "AST.CASTELLANE.JET")

asset("AST.CASTELLANE.AUTO", "12-Car Hypercar Collection", "auto", "PRIN.CASTELLANE", "Curated Collection", "N/A",
      ["COV.CASTAUTO.FLEET", "COV.CASTAUTO.TRANSIT"])
coverage("COV.CASTAUTO.FLEET", "Agreed Value Fleet", "auto", "agreed_value_fleet", "AST.CASTELLANE.AUTO")
coverage("COV.CASTAUTO.TRANSIT", "Transit", "auto", "transit", "AST.CASTELLANE.AUTO")

asset("AST.CASTELLANE.PERSONAL", "The Family / Board", "personal", "PRIN.CASTELLANE", "Executive Risk Program", "N/A",
      ["COV.CASTPER.KR", "COV.CASTPER.DO", "COV.CASTPER.CYBER"])
coverage("COV.CASTPER.KR", "Kidnap & Ransom (K&R)", "personal", "kr", "AST.CASTELLANE.PERSONAL")
coverage("COV.CASTPER.DO", "Executive D&O", "personal", "exec_do", "AST.CASTELLANE.PERSONAL")
coverage("COV.CASTPER.CYBER", "Cyber Crime", "personal", "cyber_crime", "AST.CASTELLANE.PERSONAL")

asset("AST.BELLWEATHER.HELI", "Sikorsky S-76", "aviation", "PRIN.BELLWEATHER", "Helicopter", "N/A",
      ["COV.BWHELI.HULL", "COV.BWHELI.LIAB"])
coverage("COV.BWHELI.HULL", "Aviation Hull", "aviation", "aviation_hull", "AST.BELLWEATHER.HELI")
coverage("COV.BWHELI.LIAB", "Aviation Liability", "aviation", "aviation_liability", "AST.BELLWEATHER.HELI")

asset("AST.OKONKWO.YACHT", "M/Y Meridian II, 38m", "yacht", "PRIN.OKONKWO", "38m Yacht", "S4",
      ["COV.OKONKWO.HULL", "COV.OKONKWO.PI"])
coverage("COV.OKONKWO.HULL", "Hull & Machinery", "yacht", "hull_machinery", "AST.OKONKWO.YACHT")
coverage("COV.OKONKWO.PI", "P&I", "yacht", "pi", "AST.OKONKWO.YACHT")

# ============================================================
# EDGES_720
# ============================================================
edges = []

def edge(id_, edge_type, from_, from_v, to, to_v, why, freq=None, via=None, narrative=None):
    e = {"id": id_, "edge_type": edge_type, "from": from_, "from_vertical": from_v,
         "to": to, "to_vertical": to_v, "why": why}
    if freq: e["freq"] = freq
    if via: e["via_relationship"] = via
    if narrative: e["generated_narrative"] = narrative
    edges.append(e)

# structural "knows" edges (in-vertical, real ontology-style relationships)
edge("E.001", "knows", "OWN.9", "yacht", "OWN.12", "yacht",
     "SPV Director reports to Trustee as sole shareholder of the SPV.", "transaction-triggered")
edge("E.002", "knows", "OWN.12", "yacht", "OWN.7", "yacht",
     "Trust deed compliance: GC opinion required on every material transaction.", "transaction-triggered")
edge("E.003", "knows", "MCR.2.WHISPER", "yacht", "OWN.9", "yacht",
     "Captain briefs SPV Director on vessel status and operational matters.", "weekly")
edge("E.004", "knows", "OWN.7.MERIDIAN", "yacht", "MCR.2.SOLSTICE", "yacht",
     "Trustee receives operational briefs from Captain — has gone quiet for 8 months.", "should be weekly, now silent")
edge("E.005", "knows", "OWN.9.VOSS", "yacht", "RE.PM.VOSS", "real_estate",
     "Same family trust structure — SPV Director oversees both the yacht and the RE property manager's reporting line.", "quarterly")

# cross-pollination edges (the flagship mechanic)
edge("E.101", "cross_pollination", "OWN.7", "yacht", "COV.WHISPERJET.LIAB", "aviation",
     "Same individual (Sarah Kellerman) is Ring-1 GC for the yacht's Trust AND the Aviation program's risk buyer of record for the Gulfstream.",
     via={"ring": 1, "coi_weight": 9},
     narrative="Because Alex has a Ring 1 relationship with the Yacht's General Counsel, we have a direct, warm path to capture the $250,000 Aviation Liability policy on their Gulfstream.")
edge("E.102", "cross_pollination", "OWN.9.VOSS", "yacht", "COV.CHALET.PROPERTY", "real_estate",
     "Isabelle Voss (yacht SPV Director) also sits on the family trust entity that owns the Aspen chalet — same warm path, unexploited by Cristian to date.",
     via={"ring": 1, "coi_weight": 8},
     narrative="Because Cristian already has a Ring 1 relationship on the Voss yacht SPV, the same relationship is a direct path to the untouched $190,000 High-Value Property program on their Aspen chalet.")
edge("E.103", "cross_pollination", "AVI.RISK.CASTELLANE", "aviation", "COV.CASTPER.CYBER", "personal",
     "Marcus Delacroix (Aviation risk buyer) also sits on the Castellane Trust's personal-risk committee.",
     via={"ring": 1, "coi_weight": 8},
     narrative="Because Alex has a Ring 1 relationship with Castellane's Aviation risk buyer, the same relationship reaches the unbound $95,000 Cyber Crime line on their personal risk program.")

# ============================================================
# BOUND-POLICY OVERLAY (Epic mock) — keyed by coverage_id
# ============================================================
bound = []

def bp(coverage_id, status, responsible_node_id, bound_premium=None, carrier=None, gwp_estimate=None,
       gwp_method=None, unbound_reason=None, evidence_note=None, expiry_date=None, producer_of_record=None):
    bound.append({
        "coverage_id": coverage_id, "status": status, "responsible_node_id": responsible_node_id,
        "bound_premium": bound_premium, "carrier": carrier, "gwp_estimate": gwp_estimate,
        "gwp_method": gwp_method, "unbound_reason": unbound_reason, "evidence_note": evidence_note,
        "expiry_date": expiry_date, "producer_of_record": producer_of_record,
    })

# --- Whisper (Alex — "The Sniper," flagship cross-pollination case) ---
# Yacht coverages: OWN.9 (SPV Director) is signatory for Hull/PI/FineArt; crew-adjacent lines go to the Captain.
bp("COV.WHISPER.HULL", "bound", "OWN.9", 310000, "Example Underwriters Ltd", expiry_date="2027-02-01", producer_of_record="Alex")
bp("COV.WHISPER.PI", "bound", "OWN.9", 96000, "Example Underwriters Ltd", expiry_date="2027-02-01", producer_of_record="Alex")
bp("COV.WHISPER.CREWMED", "bound", "MCR.2.WHISPER", 41800, "Example Underwriters Ltd", expiry_date="2027-02-01", producer_of_record="Alex")
bp("COV.WHISPER.FINEART", "unbound", "MCR.2.WHISPER", gwp_estimate=20225, gwp_method="ghost_policy", unbound_reason="lapsed_with_us",
   evidence_note="Bound with us 2022-2024 at $18,000, allowed to lapse at 2024 renewal.")
# Aviation coverages: OWN.7 (Sarah Kellerman) is the dual-hat GC/Aviation risk buyer for this family office — matches the cross_pollination edge target.
bp("COV.WHISPERJET.HULL", "bound", "OWN.7", 275000, "Example Underwriters Ltd", expiry_date="2027-04-15", producer_of_record="Alex")
bp("COV.WHISPERJET.LIAB", "unbound", "OWN.7", gwp_estimate=250000, gwp_method="peer_comp", unbound_reason="never_written")
bp("COV.WHISPERJET.CREWACC", "bound", "OWN.7", 18500, "Example Underwriters Ltd", expiry_date="2027-04-15", producer_of_record="Alex")
bp("COV.WHISPERJET.WAR", "unbound", "OWN.7", gwp_estimate=34000, gwp_method="peer_comp", unbound_reason="competitor_held",
   evidence_note="Salesforce call note (2026-06-02, Alex): client mentioned current War Risk coverage through Global Aerospace — displacement opportunity, NOT unpriced risk.")

# --- Solstice (Jacob — "The Legacy," churn/expiring case) ---
bp("COV.SOLSTICE.HULL", "expiring", "OWN.7.MERIDIAN", 198000, "Northbridge Marine", expiry_date="2026-10-01", producer_of_record="Jacob")
bp("COV.SOLSTICE.PI", "expiring", "OWN.7.MERIDIAN", 71000, "Northbridge Marine", expiry_date="2026-10-01", producer_of_record="Jacob")
bp("COV.SOLSTICE.CREWMED", "expiring", "MCR.2.SOLSTICE", 30600, "Northbridge Marine", expiry_date="2026-10-01", producer_of_record="Jacob")

# --- Aurelia (Rahul — "The Grinder," high touch, weak Ring-1 penetration) ---
bp("COV.AURELIA.HULL", "bound", "OWN.9.HARTWELL", 240000, "Northbridge Marine", expiry_date="2027-01-15", producer_of_record="Rahul")
bp("COV.AURELIA.PI", "bound", "OWN.9.HARTWELL", 82000, "Northbridge Marine", expiry_date="2027-01-15", producer_of_record="Rahul")
bp("COV.AURELIA.CREWMED", "unbound", "MCR.2.AURELIA", gwp_estimate=33400, gwp_method="parameter_multiplier", unbound_reason="never_written")
bp("COV.AURELIA.FINEART", "unbound", "OWN.9.HARTWELL", gwp_estimate=24800, gwp_method="peer_comp", unbound_reason="never_written")

# --- Halcyon + Chalet (Cristian — "The Silo," single-vertical only) ---
bp("COV.HALCYON.HULL", "bound", "OWN.9.VOSS", 165000, "Example Underwriters Ltd", expiry_date="2027-03-01", producer_of_record="Cristian")
bp("COV.HALCYON.PI", "bound", "OWN.9.VOSS", 58000, "Example Underwriters Ltd", expiry_date="2027-03-01", producer_of_record="Cristian")
# Chalet: WINDFLOOD/LIAB go to the native RE role (no cross-pollination edge to them); PROPERTY is the specific
# coverage the E.102 cross_pollination edge targets, so its responsible node is the yacht-side path (OWN.9.VOSS),
# not the native RE role — this is the point of the cross-pollination mechanic, not an inconsistency.
bp("COV.CHALET.WINDFLOOD", "unbound", "RE.PM.VOSS", gwp_estimate=87000, gwp_method="peer_comp", unbound_reason="never_written")
bp("COV.CHALET.PROPERTY", "unbound", "OWN.9.VOSS", gwp_estimate=190000, gwp_method="peer_comp", unbound_reason="never_written")
bp("COV.CHALET.LIAB", "unbound", "RE.PM.VOSS", gwp_estimate=41000, gwp_method="peer_comp", unbound_reason="never_written")

# --- Castellane (Alex — 2nd account, multi-vertical, feeds the Aviation Liability cluster + personal cross-pollination) ---
bp("COV.CASTJET.HULL", "bound", "AVI.RISK.CASTELLANE", 260000, "Example Underwriters Ltd", expiry_date="2027-05-01", producer_of_record="Alex")
bp("COV.CASTJET.LIAB", "unbound", "AVI.RISK.CASTELLANE", gwp_estimate=245000, gwp_method="peer_comp", unbound_reason="never_written")
bp("COV.CASTJET.CREWACC", "bound", "AVI.RISK.CASTELLANE", 17200, "Example Underwriters Ltd", expiry_date="2027-05-01", producer_of_record="Alex")
bp("COV.CASTAUTO.FLEET", "bound", "AUTO.MGR.CASTELLANE", 88000, "Northbridge Marine", expiry_date="2027-06-01", producer_of_record="Alex")
bp("COV.CASTAUTO.TRANSIT", "unbound", "AUTO.MGR.CASTELLANE", gwp_estimate=12000, gwp_method="parameter_multiplier", unbound_reason="never_written")
bp("COV.CASTPER.KR", "bound", "PER.EA.CASTELLANE", 54000, "Example Underwriters Ltd", expiry_date="2027-06-01", producer_of_record="Alex")
bp("COV.CASTPER.DO", "unbound", "PER.EA.CASTELLANE", gwp_estimate=61000, gwp_method="peer_comp", unbound_reason="never_written")
# CYBER is the specific coverage E.103's cross_pollination edge targets, so its responsible node is the aviation-side path.
bp("COV.CASTPER.CYBER", "unbound", "AVI.RISK.CASTELLANE", gwp_estimate=95000, gwp_method="peer_comp", unbound_reason="never_written")

# --- Bellweather (Jacob — 3rd Aviation Liability cluster member) ---
bp("COV.BWHELI.HULL", "bound", "AVI.RISK.BELLWEATHER", 145000, "Example Underwriters Ltd", expiry_date="2027-07-01", producer_of_record="Jacob")
bp("COV.BWHELI.LIAB", "unbound", "AVI.RISK.BELLWEATHER", gwp_estimate=118000, gwp_method="peer_comp", unbound_reason="never_written")

# --- Okonkwo (Rahul — fully lapsed / Lost account) ---
bp("COV.OKONKWO.HULL", "unbound", "OWN.9.OKONKWO", gwp_estimate=142000, gwp_method="ghost_policy", unbound_reason="lapsed_with_us",
   evidence_note="Entire account lapsed 14 months ago — full win-back, not a line-item upsell.")
bp("COV.OKONKWO.PI", "unbound", "OWN.9.OKONKWO", gwp_estimate=48000, gwp_method="ghost_policy", unbound_reason="lapsed_with_us",
   evidence_note="Entire account lapsed 14 months ago — full win-back, not a line-item upsell.")

# ============================================================
# PRODUCER ROSTER (staged names as instructed) + 90-day-ago snapshot for Comet Tail
# ============================================================
producers = [
    {"producer_id": "REP.ALEX", "name": "Alex", "archetype": "High-Yield, Low-Touch", "region": "South Florida",
     "narrative": "80% Ring 1/2 signatories, massive GWP. The ideal state — talks to fewer people, exclusively economic buyers.",
     "touchpoints_90d": 22, "gwp_captured_90d": 1850000, "pct_touches_ring1_2": 0.80,
     "touchpoints_90d_prior": 26, "gwp_captured_90d_prior": 1510000, "churn_risk_score": 1.2},
    {"producer_id": "REP.RAHUL", "name": "Rahul", "archetype": "High-Touch, Low-Yield", "region": "South Florida",
     "narrative": "90% Ring 3/4 operational/crew, low GWP. Massive activity, almost none of it lands on a signatory.",
     "touchpoints_90d": 210, "gwp_captured_90d": 240000, "pct_touches_ring1_2": 0.10,
     "touchpoints_90d_prior": 188, "gwp_captured_90d_prior": 255000, "churn_risk_score": 2.0},
    {"producer_id": "REP.JACOB", "name": "Jacob", "archetype": "Legacy Book, Decaying Contact", "region": "South Florida",
     "narrative": "Massive GWP, but comm-frequency is decaying. Bound large accounts years ago, hasn't spoken to the Ring 1 signatory in 8 months.",
     "touchpoints_90d": 9, "gwp_captured_90d": 2100000, "pct_touches_ring1_2": 0.35,
     "touchpoints_90d_prior": 21, "gwp_captured_90d_prior": 2140000, "churn_risk_score": 8.6},
    {"producer_id": "REP.CRISTIAN", "name": "Cristian", "archetype": "Single-Vertical Concentration", "region": "South Florida",
     "narrative": "100% single-vertical GWP, zero adjacent policies. Has the Ring 1 ear of the UBO, leaves the cross-pollination money on the table.",
     "touchpoints_90d": 60, "gwp_captured_90d": 680000, "pct_touches_ring1_2": 0.75,
     "touchpoints_90d_prior": 58, "gwp_captured_90d_prior": 665000, "churn_risk_score": 1.8},
]

leadership = [
    {"leader_id": "LEAD.JONATHAN", "name": "Jonathan", "title": "Regional Director", "scope": "South Florida region — team lead tier, sees all 4 producers"},
    {"leader_id": "LEAD.TRISH", "name": "Trish", "title": "SVP", "scope": "Executive — sees the full regional roll-up"},
]

account_producer_map = [
    {"ubo_id": "PRIN.MARCHETTI", "producer_id": "REP.ALEX"},
    {"ubo_id": "PRIN.MERIDIAN", "producer_id": "REP.JACOB"},
    {"ubo_id": "PRIN.HARTWELL", "producer_id": "REP.RAHUL"},
    {"ubo_id": "PRIN.VOSS", "producer_id": "REP.CRISTIAN"},
    {"ubo_id": "PRIN.CASTELLANE", "producer_id": "REP.ALEX"},
    {"ubo_id": "PRIN.BELLWEATHER", "producer_id": "REP.JACOB"},
    {"ubo_id": "PRIN.OKONKWO", "producer_id": "REP.RAHUL"},
]

# ============================================================
# PRODUCER OVERLAY — relationship weight (per producer x node) + referral edges
# ============================================================
relationship_weights = [
    # Alex — Whisper + Castellane (known, strong on signatories)
    {"producer_id": "REP.ALEX", "node_id": "OWN.9", "relationship_weight": 8.7, "activity_count_90d": 14, "last_activity_date": "2026-08-10", "contact_status": "active"},
    {"producer_id": "REP.ALEX", "node_id": "MCR.2.WHISPER", "relationship_weight": 5.1, "activity_count_90d": 6, "last_activity_date": "2026-08-05", "contact_status": "active"},
    {"producer_id": "REP.ALEX", "node_id": "AVI.RISK.CASTELLANE", "relationship_weight": 8.2, "activity_count_90d": 11, "last_activity_date": "2026-08-09", "contact_status": "active"},
    {"producer_id": "REP.ALEX", "node_id": "PER.EA.CASTELLANE", "relationship_weight": 6.4, "activity_count_90d": 7, "last_activity_date": "2026-07-28", "contact_status": "active"},
    # OWN.7 (GC) and OWN.12 (Trustee) deliberately have NO row for Alex — untapped/grey, per the Whisper cross-pollination story
    # Rahul — Aurelia (all Ring 3/4, zero Ring-1 rows = the Grinder story)
    {"producer_id": "REP.RAHUL", "node_id": "MCR.2.AURELIA", "relationship_weight": 6.8, "activity_count_90d": 40, "last_activity_date": "2026-08-13", "contact_status": "active"},
    {"producer_id": "REP.RAHUL", "node_id": "MCR.4.AURELIA", "relationship_weight": 5.5, "activity_count_90d": 55, "last_activity_date": "2026-08-14", "contact_status": "active"},
    {"producer_id": "REP.RAHUL", "node_id": "MCR.6.AURELIA", "relationship_weight": 4.2, "activity_count_90d": 61, "last_activity_date": "2026-08-14", "contact_status": "active"},
    # Jacob — Solstice (Trustee has decayed to zero) + Bellweather (known)
    {"producer_id": "REP.JACOB", "node_id": "MCR.2.SOLSTICE", "relationship_weight": 4.4, "activity_count_90d": 5, "last_activity_date": "2026-07-20", "contact_status": "active"},
    {"producer_id": "REP.JACOB", "node_id": "MCR.3.SOLSTICE", "relationship_weight": 3.8, "activity_count_90d": 4, "last_activity_date": "2026-07-18", "contact_status": "active"},
    {"producer_id": "REP.JACOB", "node_id": "AVI.RISK.BELLWEATHER", "relationship_weight": 7.9, "activity_count_90d": 9, "last_activity_date": "2026-08-11", "contact_status": "active"},
    # OWN.7.MERIDIAN (Trustee) deliberately has NO row for Jacob within the last 8 months — the churn signal itself
    # Cristian — Voss (strong Ring-1, single vertical only — the Silo story)
    {"producer_id": "REP.CRISTIAN", "node_id": "OWN.9.VOSS", "relationship_weight": 9.1, "activity_count_90d": 16, "last_activity_date": "2026-08-12", "contact_status": "active"},
    {"producer_id": "REP.CRISTIAN", "node_id": "OWN.12", "relationship_weight": 3.0, "activity_count_90d": 2, "last_activity_date": "2026-05-01", "contact_status": "active"},
    # RE.PM.VOSS deliberately has NO row for Cristian — untapped, despite the cross-pollination edge existing structurally
]

referral_edges = [
    {"id": "REF.001", "producer_id": "REP.ALEX", "from_node_id": "OWN.9", "to_node_id": "MCR.2.WHISPER", "referred_date": "2026-03-14", "notes": "SPV Director introduced Alex to the Captain at haul-out."},
    {"id": "REF.002", "producer_id": "REP.CRISTIAN", "from_node_id": "OWN.9.VOSS", "to_node_id": "OWN.12", "referred_date": "2026-01-09", "notes": "SPV Director introduced Cristian to the family Trustee."},
]

# ============================================================
# CATEGORY — "industry type" segmentation for Dashboard 1's wheel
# (distinct from `vertical`: category groups by FUNCTION across verticals,
# e.g. every Captain/Chief Engineer regardless of which yacht, so the wheel
# reads as 8 industry-type wedges instead of 1-2 account wedges.)
# ============================================================
CATEGORY_MAP = {
    "PRIN.MARCHETTI": "Ownership & Family Office", "PRIN.MERIDIAN": "Ownership & Family Office",
    "PRIN.HARTWELL": "Ownership & Family Office", "PRIN.VOSS": "Ownership & Family Office",
    "PRIN.CASTELLANE": "Ownership & Family Office", "PRIN.BELLWEATHER": "Ownership & Family Office",
    "PRIN.OKONKWO": "Ownership & Family Office",
    "OWN.7": "Legal & Finance", "OWN.9": "Ownership & Family Office", "OWN.12": "Ownership & Family Office",
    "OWN.7.MERIDIAN": "Ownership & Family Office", "OWN.9.HARTWELL": "Ownership & Family Office",
    "OWN.9.VOSS": "Ownership & Family Office", "OWN.9.OKONKWO": "Ownership & Family Office",
    "MCR.2.WHISPER": "Marine Crew & Command", "MCR.2.SOLSTICE": "Marine Crew & Command",
    "MCR.3.SOLSTICE": "Marine Crew & Command", "MCR.2.AURELIA": "Marine Crew & Command",
    "MCR.4.AURELIA": "Marine Crew & Command", "MCR.6.AURELIA": "Marine Crew & Command",
    "AVI.RISK.CASTELLANE": "Aviation Ops", "AVI.PILOT.CASTELLANE": "Aviation Ops",
    "AVI.RISK.BELLWEATHER": "Aviation Ops", "AVI.PILOT.BELLWEATHER": "Aviation Ops",
    "RE.PM.VOSS": "Real Estate & Property",
    "AUTO.MGR.CASTELLANE": "Automotive & Collections",
    "PER.EA.CASTELLANE": "Personal & Executive Staff",
}
for n in nodes:
    if n["node_type"] in ("role", "principal"):
        n["category"] = CATEGORY_MAP.get(n["id"], "Ownership & Family Office")

# ============================================================
# PROSPECT POOL — 100 per producer (75 known / 25 untapped), staged/fake,
# to make the whole-book wheel read as a real Salesforce contact list rather
# than the ~8-person subset tied to the flagship named deal accounts.
# ============================================================
import random
random.seed(720)

FIRST_NAMES = ["James","Olivia","William","Sophia","Henry","Amelia","Charles","Isla","Edward","Grace",
    "Theodore","Charlotte","Arthur","Eleanor","Frederick","Beatrice","Julian","Margot","Sebastian","Clara",
    "Nathaniel","Rosalind","Alexander","Vivienne","Benjamin","Cecilia","Christopher","Adelaide","Dominic","Josephine",
    "Gabriel","Marguerite","Lucas","Genevieve","Maximilian","Isabelle","Anthony","Camille","Rupert","Odette",
    "Desmond","Wilhelmina","Percival","Constance","Reginald","Theodora","Bartholomew","Anastasia","Cornelius","Seraphina"]
LAST_NAMES = ["Ashworth","Beaumont","Carrington","Delacroix","Ellsworth","Fairweather","Grosvenor","Hartley",
    "Ivanovic","Jarnac","Kensington","Lindqvist","Marchetti","Northcote","Ogilvie","Pemberton","Quintrell",
    "Ravensworth","Sinclair","Thornbury","Underhill","Vandermeer","Whitmore","Xavier","Yardley","Zamora",
    "Ashcombe","Blackwood","Chastain","Devereux","Everleigh","Fenwick","Grantham","Halloway","Isherwood",
    "Judd","Kilbride","Lachance","Moreau","Nightingale","Osgood","Prentiss","Radcliffe","Stavros","Trevelyan"]

CATEGORIES = {
    "Ownership & Family Office": {"ring": [0,1], "vertical": ["yacht","aviation","real_estate","auto","personal"],
        "titles": ["Family Office CEO","Family Office COO","Trustee","SPV Director","Protector","Chief of Staff","Executive Assistant to Principal","Holding Company Director"]},
    "Legal & Finance": {"ring": [1,2], "vertical": ["yacht","real_estate","personal"],
        "titles": ["General Counsel","Outside Maritime Counsel","Tax Advisor","Private Banker","Trust Officer","Registered Agent","Offshore Counsel"]},
    "Marine Crew & Command": {"ring": [1,2,3], "vertical": ["yacht"],
        "titles": ["Captain","Chief Engineer","Chief Officer","ETO","Purser","Chief Stewardess","Bosun","Second Engineer","Chef"]},
    "Aviation Ops": {"ring": [1,2,3], "vertical": ["aviation"],
        "titles": ["Chief Pilot","First Officer","Aviation Risk Manager","Maintenance Director","Flight Coordinator","Cabin Attendant Lead"]},
    "Real Estate & Property": {"ring": [1,2,3], "vertical": ["real_estate"],
        "titles": ["Property Manager","Estate Manager","Facilities Director","Head of Household","Landscape Director"]},
    "Automotive & Collections": {"ring": [1,2], "vertical": ["auto"],
        "titles": ["Collection Manager","Restoration Specialist","Fleet Logistics Coordinator","Concours Curator"]},
    "Personal & Executive Staff": {"ring": [1,2,3], "vertical": ["personal"],
        "titles": ["Executive Assistant","Chief of Staff","Security Director","Household Manager","Family Office Travel Lead","Personal Physician Liaison"]},
    "Insurance & Underwriting Market": {"ring": [2,3], "vertical": ["yacht","aviation","real_estate","auto","personal"],
        "titles": ["Placing Broker","Hull & Machinery Underwriter","Liability Underwriter","Marine Surveyor","Claims Adjuster","Central Agent","Facultative Reinsurance Broker"]},
}
CAT_LIST = list(CATEGORIES.keys())

# Per-producer category weighting — reflects each archetype's actual book shape.
PRODUCER_CATEGORY_WEIGHTS = {
    "REP.ALEX":     {"Ownership & Family Office": 22, "Legal & Finance": 16, "Marine Crew & Command": 6,
                      "Aviation Ops": 18, "Real Estate & Property": 4, "Automotive & Collections": 6,
                      "Personal & Executive Staff": 14, "Insurance & Underwriting Market": 14},
    "REP.RAHUL":    {"Ownership & Family Office": 6, "Legal & Finance": 4, "Marine Crew & Command": 42,
                      "Aviation Ops": 4, "Real Estate & Property": 4, "Automotive & Collections": 4,
                      "Personal & Executive Staff": 6, "Insurance & Underwriting Market": 30},
    "REP.JACOB":    {"Ownership & Family Office": 16, "Legal & Finance": 10, "Marine Crew & Command": 20,
                      "Aviation Ops": 14, "Real Estate & Property": 6, "Automotive & Collections": 4,
                      "Personal & Executive Staff": 10, "Insurance & Underwriting Market": 20},
    "REP.CRISTIAN": {"Ownership & Family Office": 26, "Legal & Finance": 8, "Marine Crew & Command": 32,
                      "Aviation Ops": 2, "Real Estate & Property": 10, "Automotive & Collections": 2,
                      "Personal & Executive Staff": 8, "Insurance & Underwriting Market": 12},
}
PRODUCER_KNOWN_RATIO = {"REP.ALEX": 0.75, "REP.RAHUL": 0.75, "REP.JACOB": 0.65, "REP.CRISTIAN": 0.72}
POOL_SIZE_PER_PRODUCER = 100

used_names = set()
def unique_name():
    while True:
        nm = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        if nm not in used_names:
            used_names.add(nm)
            return nm

pool_relationship_weights = []
for pid, weights in PRODUCER_CATEGORY_WEIGHTS.items():
    cats_weighted = []
    for cat, w in weights.items():
        cats_weighted += [cat] * w
    known_target = round(POOL_SIZE_PER_PRODUCER * PRODUCER_KNOWN_RATIO[pid])
    for i in range(POOL_SIZE_PER_PRODUCER):
        cat = random.choice(cats_weighted)
        spec = CATEGORIES[cat]
        ring = random.choice(spec["ring"])
        vertical = random.choice(spec["vertical"])
        title = random.choice(spec["titles"])
        name = unique_name()
        short = pid.split(".")[1].lower()
        node_id = f"POOL.{short}.{i:03d}"
        coi_type = "signatory" if ring <= 1 else ("conditional" if ring == 2 else "operational")
        nodes.append({
            "id": node_id, "label": f"{name} — {title}", "node_type": "role", "vertical": vertical,
            "ring": ring, "domain": 1, "coi_type": coi_type, "coi_weight": max(2, 10 - ring * 2),
            "coi_freq": random.randint(3, 10), "desc": f"{title}, staged Salesforce prospect-list contact.",
            "category": cat,
        })
        is_known = i < known_target
        if is_known:
            pool_relationship_weights.append({
                "producer_id": pid, "node_id": node_id,
                "relationship_weight": round(random.uniform(2.0, 9.5), 1),
                "activity_count_90d": random.randint(1, 25),
                "last_activity_date": f"2026-{random.randint(4,8):02d}-{random.randint(1,28):02d}",
                "contact_status": "active",
            })
        # else: no relationship_weight row at all — untapped, per the standing "absence = untapped" rule
relationship_weights += pool_relationship_weights

# ============================================================
# COVERAGE_CATALOG + DOMAINS_BY_VERTICAL
# ============================================================
coverage_catalog = {
    "yacht": [
        {"coverage_type_id": "hull_machinery", "label": "Hull & Machinery"},
        {"coverage_type_id": "pi", "label": "P&I"},
        {"coverage_type_id": "crew_medical_mlc", "label": "Crew Medical (MLC)"},
        {"coverage_type_id": "fine_art", "label": "Fine Art"},
    ],
    "aviation": [
        {"coverage_type_id": "aviation_hull", "label": "Aviation Hull"},
        {"coverage_type_id": "aviation_liability", "label": "Aviation Liability"},
        {"coverage_type_id": "crew_accident", "label": "Crew Accident"},
        {"coverage_type_id": "war_risk", "label": "War Risk"},
    ],
    "real_estate": [
        {"coverage_type_id": "coastal_wind_flood", "label": "Coastal Wind-Flood"},
        {"coverage_type_id": "high_value_property", "label": "High-Value Property"},
        {"coverage_type_id": "premises_liability", "label": "Premises Liability"},
    ],
    "auto": [
        {"coverage_type_id": "agreed_value_fleet", "label": "Agreed Value Fleet"},
        {"coverage_type_id": "transit", "label": "Transit"},
        {"coverage_type_id": "garage_keepers", "label": "Garage Keepers"},
    ],
    "personal": [
        {"coverage_type_id": "kr", "label": "K&R"},
        {"coverage_type_id": "exec_do", "label": "Executive D&O"},
        {"coverage_type_id": "cyber_crime", "label": "Cyber Crime"},
        {"coverage_type_id": "blanket_personal_articles", "label": "Blanket Personal Articles"},
    ],
}

domains_by_vertical = {
    "yacht": {"count": 9, "list": [
        {"id": 1, "name": "Ownership & Structure"}, {"id": 2, "name": "Legal & Finance"},
        {"id": 3, "name": "Flag / Class / Compliance"}, {"id": 4, "name": "Insurance Market"},
        {"id": 5, "name": "Acquisition & Sale"}, {"id": 6, "name": "New Build & Refit"},
        {"id": 7, "name": "Manning & Crew"}, {"id": 8, "name": "Charter & Guest Ops"},
        {"id": 9, "name": "Operations & Shoreside"},
    ]},
    "aviation": {"count": 4, "list": [
        {"id": 1, "name": "Ownership & Structure"}, {"id": 2, "name": "Airworthiness & Compliance"},
        {"id": 3, "name": "FBOs & Ground Ops"}, {"id": 4, "name": "Insurance & Finance"},
    ]},
    "real_estate": {"count": 3, "list": [
        {"id": 1, "name": "Ownership & Structure"}, {"id": 2, "name": "Property Management"},
        {"id": 3, "name": "Insurance & Finance"},
    ]},
    "auto": {"count": 2, "list": [
        {"id": 1, "name": "Ownership & Collection Management"}, {"id": 2, "name": "Insurance & Logistics"},
    ]},
    "personal": {"count": 2, "list": [
        {"id": 1, "name": "Family & Household"}, {"id": 2, "name": "Executive & Cyber Risk"},
    ]},
}

# ============================================================
# CONSISTENCY CHECKS — catch exactly the silent-ID-drop failure mode the audit found
# ============================================================
node_ids = {n["id"] for n in nodes}
coverage_ids = {n["id"] for n in nodes if n["node_type"] == "coverage"}
producer_ids = {p["producer_id"] for p in producers}

errors = []
for e in edges:
    if e["from"] not in node_ids: errors.append(f"edge {e['id']}: from={e['from']} not a known node")
    if e["to"] not in node_ids: errors.append(f"edge {e['id']}: to={e['to']} not a known node")
for b in bound:
    if b["coverage_id"] not in coverage_ids: errors.append(f"bound-overlay: coverage_id={b['coverage_id']} not a known coverage node")
    if b["responsible_node_id"] not in node_ids: errors.append(f"bound-overlay {b['coverage_id']}: responsible_node_id={b['responsible_node_id']} not a known node")
for rw in relationship_weights:
    if rw["node_id"] not in node_ids: errors.append(f"relationship_weight: node_id={rw['node_id']} not a known node")
    if rw["producer_id"] not in producer_ids: errors.append(f"relationship_weight: producer_id={rw['producer_id']} not in roster")
for r in referral_edges:
    if r["from_node_id"] not in node_ids: errors.append(f"referral {r['id']}: from_node_id not known")
    if r["to_node_id"] not in node_ids: errors.append(f"referral {r['id']}: to_node_id not known")
    if r["producer_id"] not in producer_ids: errors.append(f"referral {r['id']}: producer_id not in roster")
for m in account_producer_map:
    if m["ubo_id"] not in node_ids: errors.append(f"account_producer_map: ubo_id={m['ubo_id']} not known")
    if m["producer_id"] not in producer_ids: errors.append(f"account_producer_map: producer_id not in roster")
# every asset's required_coverage_ids must exist as real coverage nodes, and every coverage's parent_asset_ref must exist
for n in nodes:
    if n["node_type"] == "asset":
        for cid in n["required_coverage_ids"]:
            if cid not in coverage_ids: errors.append(f"asset {n['id']}: required_coverage_ids has unknown {cid}")
    if n["node_type"] == "coverage":
        if n["parent_asset_ref"] not in node_ids: errors.append(f"coverage {n['id']}: parent_asset_ref {n['parent_asset_ref']} unknown")
# every coverage should have exactly one bound-overlay row
bound_coverage_ids = {b["coverage_id"] for b in bound}
missing_overlay = coverage_ids - bound_coverage_ids
if missing_overlay: errors.append(f"coverages with NO bound-policy-overlay row: {missing_overlay}")

if errors:
    print("CONSISTENCY ERRORS FOUND:")
    for e in errors: print(" -", e)
    raise SystemExit(1)
print("Consistency check: PASSED —", len(nodes), "nodes,", len(edges), "edges,", len(bound), "bound-overlay rows,",
      len(relationship_weights), "relationship-weight rows,", len(producers), "producers.")

# ============================================================
# WRITE OUT AS EMBEDDABLE JS (script-tag friendly, no fetch/CORS issues under file://)
# ============================================================
def js(varname, value):
    return f"const {varname} = {json.dumps(value, indent=2)};\n"

with open(f"{OUT_DIR}/ontology-720.js", "w") as f:
    f.write("// Influence 720 — shared ontology. Generated dataset, staged/fake for illustration only.\n")
    f.write(js("ONTOLOGY_NODES", nodes))
    f.write(js("EDGES_720", edges))
    f.write(js("COVERAGE_CATALOG", coverage_catalog))
    f.write(js("DOMAINS_BY_VERTICAL", domains_by_vertical))

with open(f"{OUT_DIR}/overlays.js", "w") as f:
    f.write("// Influence 720 — Salesforce/Epic mock overlays. Staged/fake for illustration only.\n")
    f.write(js("BOUND_POLICY_OVERLAY", bound))
    f.write(js("RELATIONSHIP_WEIGHTS", relationship_weights))
    f.write(js("REFERRAL_EDGES", referral_edges))

with open(f"{OUT_DIR}/rosters.js", "w") as f:
    f.write("// Influence 720 — producer/leadership rosters. Staged/fake for illustration only.\n")
    f.write(js("PRODUCERS", producers))
    f.write(js("LEADERSHIP", leadership))
    f.write(js("ACCOUNT_PRODUCER_MAP", account_producer_map))

print("Wrote ontology-720.js, overlays.js, rosters.js to", OUT_DIR)

# ============================================================
# EXCEL EXPORTS — computed directly from the SAME in-memory structures above.
# Single source of truth: these can never drift from the HTML build, because
# nothing here is hand-typed independently — it's all derived from nodes/edges/
# bound/producers/etc. Replaces the old hand-authored drafts entirely.
# ============================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_OUT = "/Users/celeste7/Documents/INFLUENCE/02-dashboards"
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SCHEMA_FILL = PatternFill(start_color="27B3E3", end_color="27B3E3", fill_type="solid")
NOTE_FONT = Font(italic=True, color="8A8A8A")

node_by_id = {n["id"]: n for n in nodes}
producer_name_by_id = {p["producer_id"]: p["name"] for p in producers}

def lbl(node_id):
    n = node_by_id.get(node_id)
    return n["label"] if n else node_id

def write_sheet(wb, name, headers, rows, header_fill=HEADER_FILL, note=None):
    ws = wb.create_sheet(name)
    start_row = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = NOTE_FONT
        start_row = 2
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    for c, h in enumerate(headers, start=1):
        maxlen = max([len(str(h))] + [len(str(row[c-1])) for row in rows] + [10])
        ws.column_dimensions[get_column_letter(c)].width = min(maxlen + 2, 60)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return ws

GENERATED_NOTE = ("GENERATED — single source of truth is generator/gen_720_dataset.py, which also emits "
                   "the live HTML dashboards' data files. Do not hand-edit this sheet; re-run the generator.")

# ---- Dashboard 1: Producer Tactical Map ----
wb1 = openpyxl.Workbook(); wb1.remove(wb1.active)
write_sheet(wb1, "Schema", ["Entity", "Field", "Description"], [
    ["Contact (node)", "id / label / node_type / vertical / ring / domain / coi_type", "Ontology-derived, identical for every producer"],
    ["Relationship_Weight", "producer_id / node_id / relationship_weight / activity_count_90d / last_activity_date / contact_status", "Personal overlay — absence of a row means untapped/grey"],
    ["Knows_Edge", "from / to / why / freq", "Structural, ontology-derived, edge_type=knows"],
    ["Referred_Edge", "producer_id / from_node_id / to_node_id / referred_date / notes", "Historical, Salesforce-sourced"],
], header_fill=SCHEMA_FILL)
write_sheet(wb1, "Contacts_ONTOLOGY", ["id", "label", "node_type", "vertical", "ring", "domain", "coi_type"],
    [[n["id"], n["label"], n["node_type"], n["vertical"], n.get("ring"), n.get("domain"), n.get("coi_type")]
     for n in nodes if n["node_type"] in ("role", "principal")], note=GENERATED_NOTE)
write_sheet(wb1, "Relationship_Weight", ["producer_id", "producer_name", "node_id", "node_label", "relationship_weight", "activity_count_90d", "last_activity_date", "contact_status"],
    [[r["producer_id"], producer_name_by_id[r["producer_id"]], r["node_id"], lbl(r["node_id"]), r["relationship_weight"], r["activity_count_90d"], r["last_activity_date"], r["contact_status"]]
     for r in relationship_weights], note=GENERATED_NOTE)
write_sheet(wb1, "Knows_Edges", ["from", "from_label", "to", "to_label", "why", "freq"],
    [[e["from"], lbl(e["from"]), e["to"], lbl(e["to"]), e["why"], e.get("freq")]
     for e in edges if e["edge_type"] == "knows"], note=GENERATED_NOTE)
write_sheet(wb1, "Referred_Edges", ["id", "producer_id", "producer_name", "from_node_id", "from_label", "to_node_id", "to_label", "referred_date", "notes"],
    [[r["id"], r["producer_id"], producer_name_by_id[r["producer_id"]], r["from_node_id"], lbl(r["from_node_id"]), r["to_node_id"], lbl(r["to_node_id"]), r["referred_date"], r["notes"]]
     for r in referral_edges], note=GENERATED_NOTE)
wb1.save(f"{XLSX_OUT}/Dashboard1_ProducerTacticalMap.xlsx")

# ---- Dashboard 2: Whitespace / Gap Engine ----
wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)
write_sheet(wb2, "Schema", ["Entity", "Field", "Description"], [
    ["Asset", "id / label / vertical / owner_ref / asset_class / size_class", "Owned by exactly one principal"],
    ["Coverage", "id / label / vertical / coverage_type_id / parent_asset_ref", "Structural requirement — the Perfect State"],
    ["Bound_Policy_Overlay", "coverage_id / status / responsible_node_id / bound_premium / gwp_estimate / gwp_method / unbound_reason / evidence_note", "status is bound/unbound/expiring — NEVER 'gap'"],
], header_fill=SCHEMA_FILL)
write_sheet(wb2, "Assets_ONTOLOGY", ["id", "label", "vertical", "owner_ref", "owner_label", "asset_class", "size_class"],
    [[n["id"], n["label"], n["vertical"], n["owner_ref"], lbl(n["owner_ref"]), n["asset_class"], n["size_class"]]
     for n in nodes if n["node_type"] == "asset"], note=GENERATED_NOTE)
write_sheet(wb2, "Coverage_ONTOLOGY", ["id", "label", "vertical", "coverage_type_id", "parent_asset_ref", "parent_asset_label"],
    [[n["id"], n["label"], n["vertical"], n["coverage_type_id"], n["parent_asset_ref"], lbl(n["parent_asset_ref"])]
     for n in nodes if n["node_type"] == "coverage"], note=GENERATED_NOTE)
write_sheet(wb2, "Bound_Policy_Overlay",
    ["coverage_id", "coverage_label", "status", "responsible_node_id", "responsible_node_label", "bound_premium", "carrier", "gwp_estimate", "gwp_method", "unbound_reason", "evidence_note", "expiry_date", "producer_of_record"],
    [[b["coverage_id"], lbl(b["coverage_id"]), b["status"], b["responsible_node_id"], lbl(b["responsible_node_id"]),
      b["bound_premium"], b["carrier"], b["gwp_estimate"], b["gwp_method"], b["unbound_reason"], b["evidence_note"], b["expiry_date"], b["producer_of_record"]]
     for b in bound], note=GENERATED_NOTE)
write_sheet(wb2, "Cross_Pollination_Edges", ["from", "from_label", "to", "to_label", "why", "generated_narrative"],
    [[e["from"], lbl(e["from"]), e["to"], lbl(e["to"]), e["why"], e.get("generated_narrative")]
     for e in edges if e["edge_type"] == "cross_pollination"], note=GENERATED_NOTE)
write_sheet(wb2, "GWP_Pricing_Engine", ["method", "trigger_scenario", "confidence"], [
    ["Ghost Policy", "Policy previously bound, now lapsed/lost", "Exact — historical fact"],
    ["Peer Comp", "Policy never written for this client", "High — market-tested against own book"],
    ["Parameter Multiplier", "Coverage tied to a fixed operational ratio", "High — algorithmic baseline"],
], header_fill=SCHEMA_FILL)
wb2.save(f"{XLSX_OUT}/Dashboard2_WhitespaceGapEngine.xlsx")

# ---- Dashboard 3: Leadership Book Rollup ----
wb3 = openpyxl.Workbook(); wb3.remove(wb3.active)
write_sheet(wb3, "Schema", ["Entity", "Field", "Description"], [
    ["Account_Producer_Map", "ubo_id / producer_id", "Which producer services which UBO — servicing fact, not ownership"],
    ["Account_Rollup", "ubo_id / total_bound / total_whitespace / gap_count / account_status", "Computed, never hand-typed"],
], header_fill=SCHEMA_FILL)
write_sheet(wb3, "Account_Producer_Map", ["ubo_id", "ubo_label", "producer_id", "producer_name"],
    [[m["ubo_id"], lbl(m["ubo_id"]), m["producer_id"], producer_name_by_id[m["producer_id"]]]
     for m in account_producer_map], note=GENERATED_NOTE)

def account_rollup_rows():
    rows = []
    assets_by_owner = {}
    for n in nodes:
        if n["node_type"] == "asset":
            assets_by_owner.setdefault(n["owner_ref"], []).append(n["id"])
    coverage_by_asset = {}
    for n in nodes:
        if n["node_type"] == "coverage":
            coverage_by_asset.setdefault(n["parent_asset_ref"], []).append(n["id"])
    bound_by_cov = {b["coverage_id"]: b for b in bound}
    for m in account_producer_map:
        ubo = m["ubo_id"]
        cov_ids = [c for a in assets_by_owner.get(ubo, []) for c in coverage_by_asset.get(a, [])]
        rows_b = [bound_by_cov[c] for c in cov_ids]
        total_bound = sum(b["bound_premium"] or 0 for b in rows_b if b["status"] in ("bound", "expiring"))
        total_whitespace = sum(b["gwp_estimate"] or 0 for b in rows_b if b["status"] == "unbound")
        gap_count = sum(1 for b in rows_b if b["status"] == "unbound")
        whole_account_lapsed = any(b["unbound_reason"] == "lapsed_with_us" and b["evidence_note"] and "entire account" in b["evidence_note"].lower() for b in rows_b)
        any_expiring = any(b["status"] == "expiring" for b in rows_b)
        status = "Lost" if whole_account_lapsed else ("Expiring" if any_expiring else "Active")
        rows.append([ubo, lbl(ubo), total_bound, total_whitespace, gap_count, status])
    return rows

write_sheet(wb3, "Account_Rollup", ["ubo_id", "ubo_label", "total_bound_premium", "total_whitespace_gwp", "gap_count", "account_status"],
    account_rollup_rows(), note=GENERATED_NOTE)
wb3.save(f"{XLSX_OUT}/Dashboard3_LeadershipBookRollup.xlsx")

# ---- Dashboard 4: Role Macro View ----
wb4 = openpyxl.Workbook(); wb4.remove(wb4.active)
write_sheet(wb4, "Schema", ["Entity", "Field", "Description"], [
    ["Role_Rollup", "role_node_id / total_bound / total_prospected / producers touching it", "New computation — not present anywhere else in the project"],
], header_fill=SCHEMA_FILL)

def role_rollup_rows():
    rows = []
    roles_touched = sorted({b["responsible_node_id"] for b in bound})
    for role_id in roles_touched:
        role_bound = [b for b in bound if b["responsible_node_id"] == role_id]
        total_bound = sum(b["bound_premium"] or 0 for b in role_bound if b["status"] in ("bound", "expiring"))
        total_prospected = sum(b["gwp_estimate"] or 0 for b in role_bound if b["status"] == "unbound")
        touching = sorted({rw["producer_id"] for rw in relationship_weights if rw["node_id"] == role_id}
                           | {b["producer_of_record"] for b in role_bound if b["producer_of_record"]})
        touching_names = ", ".join(sorted({producer_name_by_id.get(p, p) for p in touching if p}))
        rows.append([role_id, lbl(role_id), node_by_id[role_id]["vertical"], total_bound, total_prospected, touching_names])
    return rows

write_sheet(wb4, "Role_Rollup", ["role_node_id", "role_label", "vertical", "total_bound_gwp", "total_prospected_gwp", "producers_touching"],
    role_rollup_rows(), note=GENERATED_NOTE)

def vertical_rollup_rows():
    totals = {}
    for b in bound:
        v = node_by_id[b["coverage_id"]]["vertical"]
        d = totals.setdefault(v, {"bound": 0, "prospected": 0})
        if b["status"] in ("bound", "expiring"): d["bound"] += b["bound_premium"] or 0
        if b["status"] == "unbound": d["prospected"] += b["gwp_estimate"] or 0
    return [[v, d["bound"], d["prospected"], d["bound"] + d["prospected"]] for v, d in sorted(totals.items())]

write_sheet(wb4, "Vertical_Rollup", ["vertical", "total_bound_gwp", "total_prospected_gwp", "combined_gwp"],
    vertical_rollup_rows(), note=GENERATED_NOTE)
wb4.save(f"{XLSX_OUT}/Dashboard4_RoleMacroView.xlsx")

# ---- Dashboard 5: Executive Command Center ----
wb5 = openpyxl.Workbook(); wb5.remove(wb5.active)
write_sheet(wb5, "Schema", ["Entity", "Field", "Description"], [
    ["Producers", "producer_id / name / archetype / touchpoints_90d(_prior) / gwp_captured_90d(_prior) / churn_risk_score", "_prior fields drive the Comet Tail trajectory"],
    ["Churn_Radar", "producer_id / account / months_since_contact / renewal_date / flag", "Derived from Producers + Account_Rollup"],
    ["Regional_Whitespace_By_Coverage_Type", "coverage_type_id / total_gwp", "Sum of gwp_estimate grouped by coverage type across the whole book"],
], header_fill=SCHEMA_FILL)
write_sheet(wb5, "Producers",
    ["producer_id", "name", "archetype", "region", "touchpoints_90d", "gwp_captured_90d", "pct_touches_ring1_2", "touchpoints_90d_prior", "gwp_captured_90d_prior", "churn_risk_score", "narrative"],
    [[p["producer_id"], p["name"], p["archetype"], p["region"], p["touchpoints_90d"], p["gwp_captured_90d"], p["pct_touches_ring1_2"], p["touchpoints_90d_prior"], p["gwp_captured_90d_prior"], p["churn_risk_score"], p["narrative"]]
     for p in producers], note=GENERATED_NOTE)
write_sheet(wb5, "Churn_Radar", ["producer_id", "producer_name", "churn_risk_score", "flagged_account", "flagged_ring1_node", "detail"], [
    ["REP.JACOB", "Jacob", 8.6, "M/Y Solstice (Meridian Family Office)", "OWN.7.MERIDIAN — Diane Foster, Trustee",
     "No RELATIONSHIP_WEIGHTS row for Jacob against the Trustee — contact has gone silent. Hull/PI/CrewMed all status=expiring, renewal 2026-10-01."],
], note=GENERATED_NOTE + " Top churn case only — extend with the same pattern for additional flagged accounts.")

def regional_whitespace_by_type():
    totals = {}
    for b in bound:
        if b["status"] != "unbound": continue
        ctype = node_by_id[b["coverage_id"]]["coverage_type_id"]
        totals[ctype] = totals.get(ctype, 0) + (b["gwp_estimate"] or 0)
    return [[k, v] for k, v in sorted(totals.items(), key=lambda kv: -kv[1])]

write_sheet(wb5, "Regional_Whitespace_By_Type", ["coverage_type_id", "total_gwp"],
    regional_whitespace_by_type(), note=GENERATED_NOTE)
wb5.save(f"{XLSX_OUT}/Dashboard5_ExecutiveCommandCenter.xlsx")

print("Wrote 5 matching .xlsx exports to", XLSX_OUT)
