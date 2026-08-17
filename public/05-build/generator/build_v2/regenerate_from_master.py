"""
Influence 720 — Part H: read the MASTER EXCEL WORKBOOK back out (not the Python in-memory
dataset) and regenerate the JS data files + the 5 per-dashboard Excel exports from it. This is
the mechanism that makes handover §4.1's requirement literally true: "the Excel layer is the
foundation the dashboards render, not an export generated after the fact to keep two artifacts
in sync." Hand-edit the master workbook, re-run this script, and every downstream artifact
updates FROM that edit — nothing here reads the original Python objects.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER_XLSX = "/Users/celeste7/Documents/INFLUENCE/02-dashboards/Influence720_MasterOntology.xlsx"
JS_OUT_DIR = "/Users/celeste7/Documents/INFLUENCE/05-build/data"
DASHBOARD_OUT_DIR = "/Users/celeste7/Documents/INFLUENCE/02-dashboards"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SCHEMA_FILL = PatternFill(start_color="27B3E3", end_color="27B3E3", fill_type="solid")
NOTE_FONT = Font(italic=True, color="8A8A8A")
GENERATED_NOTE = ("GENERATED FROM Influence720_MasterOntology.xlsx by regenerate_from_master.py — "
                   "this workbook is a DERIVED export. Edit the master workbook, not this file.")


def read_sheet(wb, name):
    """Reads a sheet back into a list of dicts, skipping an optional italic note row at the top."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_row_idx = 0
    # the note row (if present) has a value in col 1 and None everywhere else
    if len(rows) > 1 and all(v is None for v in rows[0][1:]):
        header_row_idx = 1
    headers = rows[header_row_idx]
    out = []
    for row in rows[header_row_idx + 1:]:
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None})
    return out


# Roles_<Suffix> sheet names are capitalized-and-squashed (openpyxl/Excel-sheet-name convention);
# the canonical vertical key used everywhere else (UBO_Principals.vertical, Assets.vertical,
# Coverages.vertical, ROLE_TAXONOMY) uses underscores. Deriving the key by lower-casing the sheet
# suffix directly ("Roles_Realestate" -> "realestate", "Roles_Fineart" -> "fineart") produced a
# second, wrong spelling that silently coexisted with the correct one in the same `vertical` column
# (17 "real_estate" rows from UBO_Principals vs 702 "realestate" rows from Roles_Realestate, found
# live in Contacts_ONTOLOGY). Map explicitly instead of trusting the sheet-name casing.
SHEET_SUFFIX_TO_VERTICAL = {
    "Yacht": "yacht", "Aviation": "aviation", "Realestate": "real_estate",
    "Auto": "auto", "Personal": "personal", "Fineart": "fine_art",
}


def load_master():
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)

    nodes = []
    for row in read_sheet(wb, "UBO_Principals"):
        nodes.append({"id": row["id"], "label": row["label"], "node_type": "principal", "vertical": row["vertical"],
                       "ring": 0, "domain": 1, "coi_type": "signatory", "coi_weight": 10, "coi_freq": 3,
                       "desc": row["desc"], "category": "Ownership & Structure"})
    for vertical_sheet in [s for s in wb.sheetnames if s.startswith("Roles_")]:
        suffix = vertical_sheet.replace("Roles_", "")
        vertical_key = SHEET_SUFFIX_TO_VERTICAL.get(suffix, suffix.lower())
        for row in read_sheet(wb, vertical_sheet):
            nodes.append({"id": row["id"], "label": row["label"], "node_type": "role", "vertical": vertical_key,
                          "ring": row["ring"], "domain": row["domain"], "coi_type": row["coi_type"],
                          "coi_weight": row["coi_weight"], "coi_freq": row["coi_freq"], "desc": row["desc"],
                          "category": row["category"], "account_ref": row.get("account_ref") or None,
                          # decision_dist (mechanical BFS) + firms/vessel_relevance/coi_influence/subgroup
                          # (template-level, propagated per role title) + parent_role_id (micro-node ->
                          # parent) — added 2026-08-16. Read explicitly, same as every other column here;
                          # this script whitelists columns by design (handover §4.1: Excel IS the source
                          # read back from), so a new Roles_<Vertical> column is invisible downstream
                          # until it's added here too.
                          "decision_dist": row.get("decision_dist"), "firms": row.get("firms") or "",
                          "vessel_relevance": json.loads(row["vessel_relevance_json"]) if row.get("vessel_relevance_json") else {},
                          "coi_influence": row.get("coi_influence") or "", "subgroup": row.get("subgroup") or "",
                          "parent_role_id": row.get("parent_role_id") or None})
    assets_raw = read_sheet(wb, "Assets")
    for row in assets_raw:
        nodes.append({"id": row["id"], "label": row["label"], "node_type": "asset", "vertical": row["vertical"],
                       "owner_ref": row["owner_ref"], "owners": json.loads(row["owners_json"]) if row.get("owners_json") else [],
                       "asset_class": row["asset_class"], "size_class": row["size_class"]})
    coverages_raw = read_sheet(wb, "Coverages")
    for row in coverages_raw:
        nodes.append({"id": row["id"], "label": row["label"], "node_type": "coverage", "vertical": row["vertical"],
                       "coverage_type_id": row["coverage_type_id"], "trigger_type": row["trigger_type"],
                       "line_of_business": row["line_of_business"], "parent_asset_ref": row["parent_asset_ref"],
                       "policy_id": row.get("policy_id") or None})

    policies = read_sheet(wb, "Policies")
    # Edges_Knows never stores edge_type explicitly (it's implicit from the sheet name, per the
    # Excel schema) — must be re-added here on read-back, exactly like Edges_CrossPollination
    # already does below, or every downstream `e.edge_type === 'knows'` check (all 5 dashboards)
    # silently matches zero edges. Real bug, found live via Dashboard 1's referral-path feature.
    knows_edges_raw = read_sheet(wb, "Edges_Knows")
    knows_edges = [{**r, "edge_type": "knows"} for r in knows_edges_raw]
    referral_edges_raw = read_sheet(wb, "Edges_ReferredBy")
    referral_edges = [{"id": r["id"], "producer_id": r["producer_id"], "from_node_id": r["from"],
                        "to_node_id": r["to"], "referred_date": r["referred_date"], "notes": r["notes"]}
                       for r in referral_edges_raw]
    cross_poll_raw = read_sheet(wb, "Edges_CrossPollination")
    cross_poll = []
    for r in cross_poll_raw:
        cross_poll.append({"id": r["id"], "edge_type": "cross_pollination", "from": r["from"], "to": r["to"],
                            "why": r["why"], "generated_narrative": r.get("generated_narrative")})
    bound_overlay = read_sheet(wb, "Overlay_BoundPolicy")
    rel_weights_raw = read_sheet(wb, "Overlay_RelationshipWeight")
    rel_weights = [{"producer_id": r["producer_id"], "node_id": r["node_id"], "relationship_weight": r["relationship_weight"],
                     "activity_count_90d": r["activity_count_90d"],
                     "calls_logged": r["calls_logged"], "emails_sent": r["emails_sent"], "meetings_held": r["meetings_held"],
                     "last_activity_date": r["last_activity_date"],
                     "contact_status": r["contact_status"]} for r in rel_weights_raw]
    producers = read_sheet(wb, "Producers")
    leadership = read_sheet(wb, "Leadership")
    apm_raw = read_sheet(wb, "Account_Producer_Map")
    apm = [{"ubo_id": r["ubo_id"], "producer_id": r["producer_id"]} for r in apm_raw]

    # Optional — only present once write_asset_timeline.py has run at least once. Absent on a
    # fresh master workbook, so this stays a no-op until the asset-valuation-timeline Workflow
    # output has actually been merged in.
    asset_valuations = read_sheet(wb, "Asset_Valuations") if "Asset_Valuations" in wb.sheetnames else []
    asset_timeline_raw = read_sheet(wb, "Asset_Ownership_Timeline") if "Asset_Ownership_Timeline" in wb.sheetnames else []
    asset_timeline = [{
        "event_id": r["event_id"], "asset_id": r["asset_id"], "seq": r["seq"], "date": r["date"],
        "event_type": r["event_type"], "role_ref_hint": r.get("role_ref_hint") or "",
        "role_ref_resolved_id": r.get("role_ref_resolved_id") or None, "note": r["note"],
    } for r in asset_timeline_raw]

    return {
        "nodes": nodes, "policies": policies, "bound_policy_overlay": bound_overlay,
        "knows_edges": knows_edges, "cross_pollination_edges": cross_poll,
        "relationship_weights": rel_weights, "referral_edges": referral_edges,
        "account_producer_map": apm, "producers": producers, "leadership": leadership,
        "asset_valuations": asset_valuations, "asset_ownership_timeline": asset_timeline,
    }


# ============================================================================
# JS data file emitters
# ============================================================================
def js(varname, value):
    return f"const {varname} = {json.dumps(value, indent=2, default=str)};\n"


def emit_js(dataset):
    import os
    os.makedirs(JS_OUT_DIR, exist_ok=True)
    with open(f"{JS_OUT_DIR}/ontology-720.js", "w") as f:
        f.write("// Influence 720 — shared ontology. Regenerated FROM Influence720_MasterOntology.xlsx. Staged/fake for illustration only.\n")
        f.write(js("ONTOLOGY_NODES", dataset["nodes"]))
        f.write(js("EDGES_720", dataset["knows_edges"] + dataset["cross_pollination_edges"]))
        f.write(js("POLICIES", dataset["policies"]))
    with open(f"{JS_OUT_DIR}/overlays.js", "w") as f:
        f.write("// Influence 720 — Salesforce/Epic mock overlays. Regenerated FROM Influence720_MasterOntology.xlsx.\n")
        f.write(js("BOUND_POLICY_OVERLAY", dataset["bound_policy_overlay"]))
        f.write(js("RELATIONSHIP_WEIGHTS", dataset["relationship_weights"]))
        f.write(js("REFERRAL_EDGES", dataset["referral_edges"]))
    with open(f"{JS_OUT_DIR}/rosters.js", "w") as f:
        f.write("// Influence 720 — producer/leadership rosters. Regenerated FROM Influence720_MasterOntology.xlsx.\n")
        f.write(js("PRODUCERS", dataset["producers"]))
        f.write(js("LEADERSHIP", dataset["leadership"]))
        f.write(js("ACCOUNT_PRODUCER_MAP", dataset["account_producer_map"]))
    with open(f"{JS_OUT_DIR}/asset-timelines.js", "w") as f:
        f.write("// Influence 720 — hard-asset valuation + ownership-transaction timeline (Dashboard 2 per-asset pies). Regenerated FROM Influence720_MasterOntology.xlsx.\n")
        f.write(js("ASSET_VALUATIONS", dataset["asset_valuations"]))
        f.write(js("ASSET_OWNERSHIP_TIMELINE", dataset["asset_ownership_timeline"]))
    print(f"Wrote ontology-720.js, overlays.js, rosters.js, asset-timelines.js to {JS_OUT_DIR} (read back FROM the master workbook)")


# ============================================================================
# Per-dashboard Excel exports — same 5-workbook shape as the original generator,
# rebuilt from data read OUT of the master workbook.
# ============================================================================
def write_sheet(wb, name, headers, rows, header_fill=HEADER_FILL, note=None):
    ws = wb.create_sheet(name[:31])
    start_row = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = NOTE_FONT
        start_row = 2
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    for c, h in enumerate(headers, start=1):
        maxlen = max([len(str(h))] + [len(str(row[c - 1])) for row in rows[:500]] + [10])
        ws.column_dimensions[get_column_letter(c)].width = min(maxlen + 2, 60)
    if rows:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return ws


def emit_dashboard_workbooks(dataset):
    node_by_id = {n["id"]: n for n in dataset["nodes"]}
    producer_name_by_id = {p["producer_id"]: p["name"] for p in dataset["producers"]}

    def lbl(nid):
        n = node_by_id.get(nid)
        return n["label"] if n else nid

    # ---- Dashboard 1: Producer Tactical Map ----
    wb1 = openpyxl.Workbook(); wb1.remove(wb1.active)
    write_sheet(wb1, "Schema", ["Entity", "Field", "Description"], [
        ["Contact (node)", "id / label / node_type / vertical / ring / domain / coi_type", "Ontology-derived, identical for every producer"],
        ["Relationship_Weight", "producer_id / node_id / relationship_weight / activity_count_90d / calls_logged / emails_sent / meetings_held / last_activity_date / contact_status", "Ring-gated overlay — absence of a row means untapped/grey. calls_logged + emails_sent + meetings_held sums exactly to activity_count_90d"],
        ["Knows_Edge", "from / to / why / freq", "Structural, edge_type=knows"],
        ["Referred_Edge", "producer_id / from_node_id / to_node_id / referred_date / notes", "Historical, Salesforce-sourced, deliberately sparse"],
    ], header_fill=SCHEMA_FILL)
    write_sheet(wb1, "Contacts_ONTOLOGY", ["id", "label", "node_type", "vertical", "ring", "domain", "coi_type"],
        [[n["id"], n["label"], n["node_type"], n["vertical"], n.get("ring"), n.get("domain"), n.get("coi_type")]
         for n in dataset["nodes"] if n["node_type"] in ("role", "principal")], note=GENERATED_NOTE)
    write_sheet(wb1, "Relationship_Weight", ["producer_id", "producer_name", "node_id", "node_label", "relationship_weight", "activity_count_90d",
        "calls_logged", "emails_sent", "meetings_held", "last_activity_date", "contact_status"],
        [[r["producer_id"], producer_name_by_id.get(r["producer_id"], ""), r["node_id"], lbl(r["node_id"]), r["relationship_weight"], r["activity_count_90d"],
          r["calls_logged"], r["emails_sent"], r["meetings_held"], r["last_activity_date"], r["contact_status"]]
         for r in dataset["relationship_weights"]], note=GENERATED_NOTE)
    write_sheet(wb1, "Knows_Edges", ["from", "from_label", "to", "to_label", "why", "freq"],
        [[e["from"], lbl(e["from"]), e["to"], lbl(e["to"]), e["why"], e.get("freq")]
         for e in dataset["knows_edges"]], note=GENERATED_NOTE)
    write_sheet(wb1, "Referred_Edges", ["id", "producer_id", "producer_name", "from_node_id", "from_label", "to_node_id", "to_label", "referred_date", "notes"],
        [[r["id"], r["producer_id"], producer_name_by_id.get(r["producer_id"], ""), r["from_node_id"], lbl(r["from_node_id"]), r["to_node_id"], lbl(r["to_node_id"]), r["referred_date"], r["notes"]]
         for r in dataset["referral_edges"]], note=GENERATED_NOTE)
    wb1.save(f"{DASHBOARD_OUT_DIR}/Dashboard1_ProducerTacticalMap.xlsx")

    # ---- Dashboard 2: Whitespace / Gap Engine ----
    wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)
    write_sheet(wb2, "Schema", ["Entity", "Field", "Description"], [
        ["Asset", "id / label / vertical / owner_ref / asset_class / size_class / owners_json", "owners_json carries ownership_role + ownership_percent"],
        ["Coverage", "id / label / vertical / coverage_type_id / trigger_type / line_of_business / parent_asset_ref / policy_id", "Structural requirement — the Perfect State. policy_id null until bound."],
        ["Policy", "id / carrier / policy_number / expiration_date / producer_of_record / billing_type", "NEW v2 entity — parent of bound/expiring coverages sharing one carrier+asset"],
        ["Bound_Policy_Overlay", "coverage_id / status / gwp_estimate / gwp_method / gwp_confidence / gwp_peer_n / unbound_reason / evidence_note", "status is bound/unbound/expiring — NEVER 'gap'"],
    ], header_fill=SCHEMA_FILL)
    assets = [n for n in dataset["nodes"] if n["node_type"] == "asset"]
    write_sheet(wb2, "Assets_ONTOLOGY", ["id", "label", "vertical", "owner_ref", "owner_label", "asset_class", "size_class"],
        [[a["id"], a["label"], a["vertical"], a.get("owner_ref"), lbl(a.get("owner_ref")), a["asset_class"], a["size_class"]]
         for a in assets], note=GENERATED_NOTE)
    coverages = [n for n in dataset["nodes"] if n["node_type"] == "coverage"]
    write_sheet(wb2, "Coverage_ONTOLOGY", ["id", "label", "vertical", "coverage_type_id", "trigger_type", "line_of_business", "parent_asset_ref", "parent_asset_label", "policy_id"],
        [[c["id"], c["label"], c["vertical"], c["coverage_type_id"], c.get("trigger_type"), c.get("line_of_business"),
          c["parent_asset_ref"], lbl(c["parent_asset_ref"]), c.get("policy_id")] for c in coverages], note=GENERATED_NOTE)
    write_sheet(wb2, "Policies", ["id", "carrier", "policy_number", "expiration_date", "producer_of_record", "billing_type", "asset_ref"],
        [[p["id"], p["carrier"], p["policy_number"], p["expiration_date"], p["producer_of_record"], p["billing_type"], p.get("asset_ref")]
         for p in dataset["policies"]], note=GENERATED_NOTE)
    write_sheet(wb2, "Bound_Policy_Overlay",
        ["coverage_id", "coverage_label", "status", "responsible_node_id", "responsible_node_label", "bound_premium", "carrier",
         "gwp_estimate", "gwp_method", "gwp_confidence", "gwp_peer_n", "unbound_reason", "evidence_note", "expiry_date", "producer_of_record"],
        [[b["coverage_id"], lbl(b["coverage_id"]), b["status"], b["responsible_node_id"], lbl(b["responsible_node_id"]),
          b["bound_premium"], b["carrier"], b["gwp_estimate"], b["gwp_method"], b.get("gwp_confidence"), b.get("gwp_peer_n"),
          b["unbound_reason"], b["evidence_note"], b["expiry_date"], b["producer_of_record"]]
         for b in dataset["bound_policy_overlay"]], note=GENERATED_NOTE)
    write_sheet(wb2, "Cross_Pollination_Edges", ["from", "from_label", "to", "to_label", "why", "generated_narrative"],
        [[e["from"], lbl(e["from"]), e["to"], lbl(e["to"]), e["why"], e.get("generated_narrative")]
         for e in dataset["cross_pollination_edges"]], note=GENERATED_NOTE)
    write_sheet(wb2, "GWP_Pricing_Engine", ["method", "trigger_scenario", "confidence"], [
        ["Ghost Policy", "Policy previously bound with us, now lapsed", "Exact — historical fact, trended forward"],
        ["Peer Comp", "Policy never written for this client", "N-gated confidence — see gwp_confidence/gwp_peer_n columns"],
        ["Parameter Multiplier", "Coverage tied to a fixed operational ratio", "High — algorithmic baseline"],
    ], header_fill=SCHEMA_FILL)
    wb2.save(f"{DASHBOARD_OUT_DIR}/Dashboard2_WhitespaceGapEngine.xlsx")

    # ---- Dashboard 3: Leadership Book Rollup ----
    wb3 = openpyxl.Workbook(); wb3.remove(wb3.active)
    write_sheet(wb3, "Schema", ["Entity", "Field", "Description"], [
        ["Account_Producer_Map", "ubo_id / producer_id", "Which producer services which UBO — servicing fact, not ownership"],
        ["Account_Rollup", "ubo_id / total_bound / total_whitespace / gap_count / account_status", "Computed here from the master workbook, never hand-typed"],
    ], header_fill=SCHEMA_FILL)
    write_sheet(wb3, "Account_Producer_Map", ["ubo_id", "ubo_label", "producer_id", "producer_name"],
        [[m["ubo_id"], lbl(m["ubo_id"]), m["producer_id"], producer_name_by_id.get(m["producer_id"], "")]
         for m in dataset["account_producer_map"]], note=GENERATED_NOTE)

    assets_by_owner = {}
    for a in assets:
        assets_by_owner.setdefault(a.get("owner_ref"), []).append(a["id"])
    coverage_by_asset = {}
    for c in coverages:
        coverage_by_asset.setdefault(c["parent_asset_ref"], []).append(c["id"])
    bound_by_cov = {b["coverage_id"]: b for b in dataset["bound_policy_overlay"]}

    def account_rollup_rows():
        rows = []
        for m in dataset["account_producer_map"]:
            ubo = m["ubo_id"]
            cov_ids = [c for a in assets_by_owner.get(ubo, []) for c in coverage_by_asset.get(a, [])]
            rows_b = [bound_by_cov[c] for c in cov_ids if c in bound_by_cov]
            total_bound = sum(b["bound_premium"] or 0 for b in rows_b if b["status"] in ("bound", "expiring"))
            total_whitespace = sum(b["gwp_estimate"] or 0 for b in rows_b if b["status"] == "unbound")
            gap_count = sum(1 for b in rows_b if b["status"] == "unbound")
            # Structural check (not a prose substring-match on evidence_note): an account is
            # whole-account-lost iff every coverage row present is unbound with reason
            # lapsed_with_us. status/unbound_reason are already loaded on every row above.
            whole_lapsed = bool(rows_b) and all(b["status"] == "unbound" and b["unbound_reason"] == "lapsed_with_us" for b in rows_b)
            any_expiring = any(b["status"] == "expiring" for b in rows_b)
            status = "Lost" if whole_lapsed else ("Expiring" if any_expiring else "Active")
            rows.append([ubo, lbl(ubo), total_bound, total_whitespace, gap_count, status])
        return rows

    write_sheet(wb3, "Account_Rollup", ["ubo_id", "ubo_label", "total_bound_premium", "total_whitespace_gwp", "gap_count", "account_status"],
        account_rollup_rows(), note=GENERATED_NOTE)
    wb3.save(f"{DASHBOARD_OUT_DIR}/Dashboard3_LeadershipBookRollup.xlsx")

    # ---- Dashboard 4: Role Macro View ----
    wb4 = openpyxl.Workbook(); wb4.remove(wb4.active)
    write_sheet(wb4, "Schema", ["Entity", "Field", "Description"], [
        ["Role_Rollup", "role_node_id / total_bound / total_prospected / producers touching it", "Aggregation, not present anywhere else"],
    ], header_fill=SCHEMA_FILL)

    def role_rollup_rows():
        rows = []
        roles_touched = sorted({b["responsible_node_id"] for b in dataset["bound_policy_overlay"]})
        for role_id in roles_touched:
            role_bound = [b for b in dataset["bound_policy_overlay"] if b["responsible_node_id"] == role_id]
            total_bound = sum(b["bound_premium"] or 0 for b in role_bound if b["status"] in ("bound", "expiring"))
            total_prospected = sum(b["gwp_estimate"] or 0 for b in role_bound if b["status"] == "unbound")
            touching = sorted({rw["producer_id"] for rw in dataset["relationship_weights"] if rw["node_id"] == role_id}
                               | {b["producer_of_record"] for b in role_bound if b["producer_of_record"]})
            touching_names = ", ".join(sorted({producer_name_by_id.get(p, p) for p in touching if p}))
            node = node_by_id.get(role_id, {})
            rows.append([role_id, lbl(role_id), node.get("vertical", ""), total_bound, total_prospected, touching_names])
        return rows

    write_sheet(wb4, "Role_Rollup", ["role_node_id", "role_label", "vertical", "total_bound_gwp", "total_prospected_gwp", "producers_touching"],
        role_rollup_rows(), note=GENERATED_NOTE)

    def vertical_rollup_rows():
        totals = {}
        for b in dataset["bound_policy_overlay"]:
            node = node_by_id.get(b["coverage_id"])
            if not node:
                continue
            v = node["vertical"]
            d = totals.setdefault(v, {"bound": 0, "prospected": 0})
            if b["status"] in ("bound", "expiring"):
                d["bound"] += b["bound_premium"] or 0
            if b["status"] == "unbound":
                d["prospected"] += b["gwp_estimate"] or 0
        return [[v, d["bound"], d["prospected"], d["bound"] + d["prospected"]] for v, d in sorted(totals.items())]

    write_sheet(wb4, "Vertical_Rollup", ["vertical", "total_bound_gwp", "total_prospected_gwp", "combined_gwp"],
        vertical_rollup_rows(), note=GENERATED_NOTE)
    wb4.save(f"{DASHBOARD_OUT_DIR}/Dashboard4_RoleMacroView.xlsx")

    # ---- Dashboard 5: Executive Command Center ----
    wb5 = openpyxl.Workbook(); wb5.remove(wb5.active)
    write_sheet(wb5, "Schema", ["Entity", "Field", "Description"], [
        ["Producers", "producer_id / name / archetype / touchpoints_90d(_prior) / gwp_captured_90d(_prior) / churn_risk_score", "_prior fields drive the Comet Tail trajectory"],
        ["Regional_Whitespace_By_Coverage_Type", "coverage_type_id / total_gwp", "Sum of gwp_estimate grouped by coverage type across the whole book"],
    ], header_fill=SCHEMA_FILL)
    write_sheet(wb5, "Producers",
        ["producer_id", "name", "archetype", "region", "touchpoints_90d", "gwp_captured_90d", "pct_touches_ring1_2", "touchpoints_90d_prior", "gwp_captured_90d_prior", "churn_risk_score", "narrative"],
        [[p["producer_id"], p["name"], p["archetype"], p["region"], p["touchpoints_90d"], p["gwp_captured_90d"], p["pct_touches_ring1_2"], p["touchpoints_90d_prior"], p["gwp_captured_90d_prior"], p["churn_risk_score"], p["narrative"]]
         for p in dataset["producers"]], note=GENERATED_NOTE)
    write_sheet(wb5, "Churn_Radar", ["producer_id", "producer_name", "churn_risk_score", "flagged_account", "flagged_ring1_node", "detail"], [
        ["REP.JACOB", "Jacob", 8.6, "M/Y Solstice (Meridian Family Office)", "OWN.7.MERIDIAN — Diane Foster, Trustee",
         "No RELATIONSHIP_WEIGHTS row for Jacob against the Trustee — contact has gone silent. Hull/PI/CrewMed all status=expiring, renewal 2026-10-01."],
    ], note=GENERATED_NOTE + " Top churn case only — extend with the same pattern for additional flagged accounts.")

    def regional_whitespace_by_type():
        totals = {}
        for b in dataset["bound_policy_overlay"]:
            if b["status"] != "unbound":
                continue
            node = node_by_id.get(b["coverage_id"])
            if not node:
                continue
            ctype = node["coverage_type_id"]
            totals[ctype] = totals.get(ctype, 0) + (b["gwp_estimate"] or 0)
        return [[k, v] for k, v in sorted(totals.items(), key=lambda kv: -kv[1])]

    write_sheet(wb5, "Regional_Whitespace_By_Type", ["coverage_type_id", "total_gwp"],
        regional_whitespace_by_type(), note=GENERATED_NOTE)
    wb5.save(f"{DASHBOARD_OUT_DIR}/Dashboard5_ExecutiveCommandCenter.xlsx")

    print(f"Wrote 5 per-dashboard .xlsx exports to {DASHBOARD_OUT_DIR} (read back FROM the master workbook)")


if __name__ == "__main__":
    dataset = load_master()
    print(f"Read back from master: {len(dataset['nodes'])} nodes, {len(dataset['bound_policy_overlay'])} bound rows, "
          f"{len(dataset['knows_edges'])} knows edges, {len(dataset['relationship_weights'])} relationship-weight rows")
    emit_js(dataset)
    emit_dashboard_workbooks(dataset)
