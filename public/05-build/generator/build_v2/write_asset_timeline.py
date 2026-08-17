"""
Writes the asset-valuation-timeline Workflow's output (research-grounded valuation_usd +
2-4 real ownership-transaction events per hard asset) into two new sheets on the master
Excel workbook: Asset_Valuations, Asset_Ownership_Timeline. Mirrors the existing
write_sheet() convention from build_master_excel.py so regenerate_from_master.py can read
these back the same way it reads every other sheet.

Run AFTER the workflow's JSON result has been saved to disk (see WORKFLOW_RESULT_PATH).
Idempotent on the SHEET level — re-running replaces both sheets rather than duplicating rows.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_PATH = "/Users/celeste7/Documents/INFLUENCE/02-dashboards/Influence720_MasterOntology.xlsx"
WORKFLOW_RESULT_PATH = "/private/tmp/claude-501/-Users-celeste7/a4eedd11-2615-4b11-a3c1-eaa58ca40bfe/scratchpad/asset_timeline_result.json"
DATASET_PATH = "/Users/celeste7/Documents/INFLUENCE/05-build/generator/build_v2/dataset_v2_final.json"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="8A8A8A")


def write_sheet(wb, name, headers, rows, note=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name[:31])
    start_row = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = NOTE_FONT
        start_row = 2
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    for c, h in enumerate(headers, start=1):
        maxlen = max([len(str(h))] + [len(str(row[c - 1])) for row in rows[:500]] + [10])
        ws.column_dimensions[get_column_letter(c)].width = min(maxlen + 2, 60)
    return ws


def build_name_index(dataset):
    """account_ref -> {person_name (lowercased): role_or_principal_id}"""
    idx = {}
    for n in dataset["nodes"]:
        if n["node_type"] == "role" and n.get("account_ref"):
            person_name = n["label"].split(" — ")[0].strip()
            idx.setdefault(n["account_ref"], {})[person_name.lower()] = n["id"]
        elif n["node_type"] == "principal":
            idx.setdefault(n["id"], {})[n["label"].lower()] = n["id"]
    return idx


def main():
    with open(WORKFLOW_RESULT_PATH) as f:
        result = json.load(f)
    with open(DATASET_PATH) as f:
        dataset = json.load(f)

    name_idx = build_name_index(dataset)

    valuation_rows = []
    timeline_rows = []
    event_seq = 0
    covered_assets = 0
    for acc in result["accounts"]:
        aid = acc["account_id"]
        acc_names = name_idx.get(aid, {})
        for asset in acc["assets"]:
            covered_assets += 1
            valuation_rows.append([asset["asset_id"], asset["valuation_usd"], asset["valuation_basis"]])
            for seq, ev in enumerate(asset.get("timeline", []), start=1):
                event_seq += 1
                hint = (ev.get("role_ref_hint") or "").strip()
                role_id = acc_names.get(hint.lower()) if hint else None
                timeline_rows.append([
                    f"EVT.{event_seq:05d}", asset["asset_id"], seq, ev["date"], ev["event_type"],
                    hint, role_id, ev["note"],
                ])

    wb = openpyxl.load_workbook(XLSX_PATH)
    write_sheet(
        wb, "Asset_Valuations", ["asset_id", "valuation_usd", "valuation_basis"], valuation_rows,
        note="GENERATED via the asset-valuation-timeline Workflow (research-grounded market bands, see methodology_note per vertical). Regenerate via write_asset_timeline.py.",
    )
    write_sheet(
        wb, "Asset_Ownership_Timeline",
        ["event_id", "asset_id", "seq", "date", "event_type", "role_ref_hint", "role_ref_resolved_id", "note"],
        timeline_rows,
        note="GENERATED — 2-4 ownership/transaction events per hard asset (yacht/aviation/real_estate/auto). role_ref_resolved_id is matched against this account's own narrative names where possible; blank means the event predates this family's ownership or involves no specific person.",
    )
    wb.save(XLSX_PATH)
    resolved = sum(1 for r in timeline_rows if r[6])
    print(f"Wrote {len(valuation_rows)} valuation rows ({covered_assets} assets covered), "
          f"{len(timeline_rows)} timeline events ({resolved} resolved to a real role/principal id)")


if __name__ == "__main__":
    main()
